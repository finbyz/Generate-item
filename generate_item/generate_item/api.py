import json
from io import BytesIO

import frappe
from frappe import _
from frappe.desk.query_report import get_column_as_dict, run
from frappe.utils import cint, cstr, flt, get_datetime
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

DATE_FIELDTYPES = {"Date"}
DATETIME_FIELDTYPES = {"Datetime"}
NUMERIC_FIELDTYPES = {"Int", "Float", "Currency", "Percent"}


@frappe.whitelist()
def export_query_report(form_params):
    """
    Export ANY Query Report / Script Report to a clean, branded Excel
    workbook - header row with filters, banded rows - and stream it
    straight back as a browser download. No File doc is created.
    """

    if isinstance(form_params, str):
        form_params = json.loads(form_params)

    if not isinstance(form_params, dict):
        frappe.throw(_("Invalid parameters."))

    form_params = frappe._dict(form_params)
    report_name = form_params.get("report_name")
    if not report_name:
        frappe.throw(_("report_name is required."))

    filters = form_params.get("filters", {})
    if isinstance(filters, str):
        try:
            filters = json.loads(filters)
        except Exception:
            filters = {}

    data = run(report_name, filters, are_default_filters=False)
    data = frappe._dict(data)

    columns = data.get("columns") or []
    rows = data.get("result") or []

    col_dicts = [get_column_as_dict(col) for col in columns]
    fieldnames = [c.get("fieldname") for c in col_dicts]
    labels = [c.get("label") or c.get("fieldname") for c in col_dicts]
    fieldtypes = [c.get("fieldtype") or "Data" for c in col_dicts]

    wb = Workbook()
    ws = wb.active
    ws.title = report_name[:31]

    # ---------------- theme ----------------
    HEADER_BLUE = "4472C4"
    WHITE = "FFFFFF"

    header_font = Font(bold=True, color=WHITE, size=10)
    header_fill = PatternFill("solid", fgColor=HEADER_BLUE)
    thin = Side(style="thin", color="B7C6E3")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    n_cols = max(len(labels), 1)
    last_col_letter = get_column_letter(n_cols)

    # ---------------- header row (row 1, no banner) ----------------
    header_row_idx = 1
    for idx, label in enumerate(labels, start=1):
        cell = ws.cell(row=header_row_idx, column=idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
    ws.row_dimensions[header_row_idx].height = 20

    max_width = [len(str(lbl)) + 4 for lbl in labels]

    # ---------------- data rows ----------------
    start_data_row = header_row_idx + 1
    for r_idx, row in enumerate(rows):
        if isinstance(row, dict):
            values = [row.get(fn, "") for fn in fieldnames]
        else:
            values = [row[i] if i < len(row) else "" for i in range(len(fieldnames))]

        excel_row_idx = start_data_row + r_idx

        for c_idx, (value, ftype) in enumerate(zip(values, fieldtypes), start=1):
            if isinstance(value, bool):
                value = "Yes" if value else "No"
            if value is None:
                value = ""

            cell = ws.cell(row=excel_row_idx, column=c_idx)

            if ftype in ("Currency", "Float") and isinstance(value, (int, float)):
                cell.value = flt(value, 2)
                cell.number_format = "#,##0.00"
                cell.alignment = right_align
            elif ftype == "Percent" and isinstance(value, (int, float)):
                cell.value = flt(value, 2) / 100
                cell.number_format = "0.00%"
                cell.alignment = right_align
            elif ftype == "Int" and isinstance(value, (int, float)):
                cell.value = cint(value)
                cell.number_format = "#,##0"
                cell.alignment = right_align
            elif ftype in DATE_FIELDTYPES and value:
                try:
                    cell.value = get_datetime(value).date()
                    cell.number_format = "dd-mm-yyyy"
                except Exception:
                    cell.value = cstr(value)
                cell.alignment = center
            elif ftype in DATETIME_FIELDTYPES and value:
                try:
                    cell.value = get_datetime(value)
                    cell.number_format = "dd-mm-yyyy hh:mm"
                except Exception:
                    cell.value = cstr(value)
                cell.alignment = center
            else:
                cell.value = value
                cell.alignment = left_align

            cell.border = thin_border

            text_value = cstr(value)
            if len(text_value) + 2 > max_width[c_idx - 1]:
                max_width[c_idx - 1] = min(len(text_value) + 2, 45)

    # ---------------- column widths ----------------
    for idx, width in enumerate(max_width, 1):
        ws.column_dimensions[get_column_letter(idx)].width = min(max(width, 10), 45)

    # ---------------- freeze header row ----------------
    ws.freeze_panes = f"A{start_data_row}"

    # ---------------- native Excel table (filters + banding) ----------------
    last_row = start_data_row + len(rows) - 1
    if rows and n_cols > 0:
        table_name = "Tbl_" + "".join(ch for ch in report_name if ch.isalnum())[:24]
        table = Table(
            displayName=table_name,
            ref=f"A{header_row_idx}:{last_col_letter}{last_row}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

        # negative numbers highlighted in red, for any numeric column
        red_font = Font(color="C00000")
        for c_idx, ftype in enumerate(fieldtypes, start=1):
            if ftype in NUMERIC_FIELDTYPES:
                col_letter = get_column_letter(c_idx)
                rng = f"{col_letter}{start_data_row}:{col_letter}{last_row}"
                ws.conditional_formatting.add(
                    rng, CellIsRule(operator="lessThan", formula=["0"], font=red_font)
                )

    # ---------------- direct download, no File doc ----------------
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = frappe.scrub(report_name) + ".xlsx"
    frappe.local.response.filename = filename
    frappe.local.response.filecontent = buffer.getvalue()
    frappe.local.response.type = "download"
    frappe.local.response.content_type = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )