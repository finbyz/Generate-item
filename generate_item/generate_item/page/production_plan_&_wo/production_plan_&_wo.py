# -*- coding: utf-8 -*-
# Copyright (c) 2026, Steelstrong / Custom App
# License: MIT / see LICENSE
"""
Production Plan & Work Order Update Control Report
====================================================

Single-screen manufacturing control center that shows, for every Sales
Order, whether a downstream change (Order Modification Request -> BOM
Modification Request -> Production Plan -> Work Order) has been fully
propagated.

Flow:
1. OMR submitted → Production Plan sales_order_modification = "Yes"
2. BMR submitted → Production Plan bom_modification = "Yes"
3. Production Plan "Get Update" button clicked → updates applied
4. Work Orders created FROM Production Plans (PP → WO relationship)
5. Report shows only WOs linked to both SO and its PPs
"""

from __future__ import unicode_literals

import json
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import frappe
from frappe import _
from frappe.utils import (
	add_days,
	add_months,
	cint,
	cstr,
	flt,
	get_first_day,
	get_last_day,
	getdate,
	nowdate,
	pretty_date, 
)


# --------------------------------------------------------------------------- #
# Badge / severity vocabulary
# --------------------------------------------------------------------------- #

BADGES = {
	"updated": {"key": "updated", "emoji": "", "label": "Updated", "color": "green"},
	"pending": {"key": "pending", "emoji": "", "label": "Pending", "color": "amber"},
	"critical": {"key": "critical", "emoji": "", "label": "Critical", "color": "red"},
	"not_required": {"key": "not_required", "emoji": "", "label": "Not Required", "color": "blue"},
	"waiting": {"key": "waiting", "emoji": "", "label": "Waiting", "color": "grey"},
}

SEVERITY_ORDER = ["Critical", "High", "Medium", "Waiting", "Low"]
SEVERITY_COLOR = {
	"Critical": "red",
	"High": "amber",
	"Medium": "blue",
	"Waiting": "grey",
	"Low": "green",
}
# Same hex values as the dashboard's CSS variables, so Excel matches the UI exactly
BADGE_FILL_HEX = {
	"updated": "5EBB63",       # green
	"pending": "F5A623",       # amber
	"critical": "E86161",      # red
	"not_required": "4C9CE2",  # blue
	"waiting": "98A6BF",       # grey
}
SEVERITY_FILL_HEX = {
	"Critical": "E86161",
	"High": "F5A623",
	"Medium": "4C9CE2",
	"Waiting": "98A6BF",
	"Low": "5EBB63",
}

STAGE_KEYS = ["so", "omr", "bmr", "pp", "wo", "completed"]
STAGE_LABEL = {
	"so": "SO",
	"omr": "OMR",
	"bmr": "BMR",
	"pp": "PP",
	"wo": "WO",
	"completed": "Completed",
}

PERIOD_OPTIONS = [
	{"value": "", "label": "All Time"},
	{"value": "today", "label": "Today"},
	{"value": "this_week", "label": "This Week"},
	{"value": "last_week", "label": "Last Week"},
	{"value": "this_month", "label": "This Month"},
	{"value": "last_month", "label": "Last Month"},
	{"value": "this_year", "label": "This Year"},
	{"value": "last_year", "label": "Last Year"},
	{"value": "custom", "label": "Custom…"},
]


def _badge(key):
	return dict(BADGES[key])


# --------------------------------------------------------------------------- #
# Public whitelisted API
# --------------------------------------------------------------------------- #


@frappe.whitelist()
def get_filter_options():
	"""Static option lists for the sticky filter bar."""
	companies = frappe.get_all("Company", pluck="name", order_by="name")
	if frappe.db.table_exists("Branch"):
		branches = frappe.get_all("Branch", pluck="name", order_by="name")
	else:
		branches = frappe.get_all(
			"Sales Order", pluck="branch", distinct=True, filters={"branch": ["is", "set"]}
		)

	return {
		"companies": companies,
		"branches": [b for b in branches if b],
		"period_options": PERIOD_OPTIONS,
		"default_from": add_days(nowdate(), -90),
		"default_to": nowdate(),
	}


@frappe.whitelist()
def get_dashboard_data(filters=None):
	"""KPI cards + overall sync ring."""
	rows = _get_computed_rows(_parse_filters(filters), limit=2000)

	pending_pp = sum(1 for r in rows if r["pp_update_required"])
	pending_wo = sum(1 for r in rows if r["wo_update_required"])
	pending_bmr = sum(1 for r in rows if r["bmr_exists"] and not r["bmr_all_approved"])
	pending_omr = sum(1 for r in rows if r["omr_exists"] and not r["omr_approved"])
	completed = sum(1 for r in rows if r["severity"] == "Low")
	critical = sum(1 for r in rows if r["severity"] == "Critical")

	trackable = sum(1 for r in rows if r["omr_exists"])
	sync_pct = round((completed / trackable) * 100, 1) if trackable else 100.0

	return {
		"pending_pp": pending_pp,
		"pending_wo": pending_wo,
		"pending_bmr": pending_bmr,
		"pending_omr": pending_omr,
		"completed": completed,
		"critical": critical,
		"total": len(rows),
		"trackable": trackable,
		"sync_pct": sync_pct,
		"all_clear": trackable > 0 and pending_pp == 0 and pending_wo == 0 and pending_bmr == 0 and critical == 0,
	}


@frappe.whitelist()
def get_grid_data(filters=None, start=0, page_length=50):
	"""Paginated grid rows."""
	start = cint(start)
	page_length = cint(page_length) or 50

	rows = _get_computed_rows(_parse_filters(filters), limit=2000)
	total = len(rows)
	page = rows[start : start + page_length]

	return {"rows": page, "total": total, "start": start, "page_length": page_length}


@frappe.whitelist()
def get_row_detail(sales_order, filters=None):
	"""Everything the right-hand drawer needs for a single Sales Order."""
	if not sales_order:
		frappe.throw(_("Sales Order is required"))

	if isinstance(filters, str):
		try:
			filters = json.loads(filters)
		except Exception:
			filters = {}

	so = frappe.db.get_value(
		"Sales Order",
		sales_order,
		[
			"name", "customer", "customer_name", "branch", "company", "transaction_date",
			"delivery_date", "status", "grand_total", "currency", "owner", "modified", "modified_by",
		],
		as_dict=True,
	)
	if not so:
		frappe.throw(_("Sales Order {0} not found").format(sales_order))

	omrs = frappe.get_all(
		"Order Modification Request",
		filters={"sales_order": sales_order},
		fields=[
			"name", "workflow_state", "docstatus", "reason_for_change", "modification_type",
			"type", "creation", "owner", "modified", "modified_by",
		],
		order_by="creation desc",
	)

	omr_names = [o.name for o in omrs]
	omr_items = []
	if omr_names:
		omr_items = frappe.get_all(
			"Sales Order Item For OMR",
			filters={"parent": ["in", omr_names]},
			fields=[
				"parent", "item", "rev_item", "description", "rev_description",
				"qty", "rev_qty", "bom_update_request", "tag_no", "line_status", "rev_line_status",
			],
		)

	bmr_names = sorted({d.bom_update_request for d in omr_items if d.bom_update_request})
	bmrs = []
	if bmr_names:
		bmrs = frappe.get_all(
			"Bom Modification Request",
			filters={"name": ["in", bmr_names]},
			fields=[
				"name", "workflow_state", "docstatus", "fg_item_code", "fg_item_name",
				"item_description", "reason_for_change", "batch_no_ref", "creation",
				"modified", "modified_by",
			],
		)

	# ------------------------------------------------------------------ #
	# FIX (Work Order count consistency): Production Plan / Work Order
	# data now comes from ONE place - _fetch_maps() - the exact same
	# helper the grid uses to build report rows. Previously this endpoint
	# ran a second, independent PP/WO lookup here (with slightly
	# different filters/joins) just to build `pps`/`wos`, while the
	# accordion counts and badges were built from a *different*
	# computation (`row`, via _fetch_maps). Two implementations of the
	# same lookup will eventually disagree - which is exactly what
	# produced a popup Work Order count that didn't match the report
	# row. There is now a single source of truth, so they cannot drift
	# apart again.
	# ------------------------------------------------------------------ #
	filtered_pp = filters.get("production_plan") if filters else None
	maps = _fetch_maps([sales_order], filtered_pp=filtered_pp)
	row = _compute_row(so, maps)

	# The one Production Plan chosen for this Sales Order (the filtered
	# PP, or - with no filter - the most recently created PP linked to
	# this SO). This is the exact same PP shown in the report row.
	chosen_pp = maps["pp"].get(sales_order)
	pps = [chosen_pp] if chosen_pp else []
	# Work Orders belonging ONLY to that chosen Production Plan - never a
	# sum across every Production Plan the Sales Order has ever had.
	wos = row["wo_list"]

	# ---- plain-English "what's still pending" list -----------------------
	pending_actions = []
	if row["omr_exists"] and not row["omr_approved"]:
		pending_actions.append(_("Get Order Modification Request {0} approved.").format(row["omr"]["name"]))
	if row["bmr_exists"] and not row["bmr_all_approved"]:
		pending_names = ", ".join(b["name"] for b in row["bmr_pending"])
		pending_actions.append(_("Approve BOM Modification Request(s): {0}.").format(pending_names))
	if row["pp_update_required"]:
		pending_actions.append(_("Click 'Get Update' on the Production Plan to apply the approved BOM change."))
	if row["wo_update_required"]:
		wo_names = ", ".join(w.get("name") for w in wos if w.get("modification_status") == "Yes")
		pending_actions.append(_("Re-issue / update Work Order(s): {0}.").format(wo_names))
	if not pending_actions:
		pending_actions.append(_("Nothing pending - fully synchronised. ✅"))

	# ---- combined chronological history -----------------------------------
	history = []
	for o in omrs:
		history.append({
			"time": o.creation, "doctype": "Order Modification Request", "name": o.name,
			"event": _("Change request raised"), "state": o.workflow_state,
			"user": o.owner,
		})
	for b in bmrs:
		history.append({
			"time": b.creation, "doctype": "Bom Modification Request", "name": b.name,
			"event": _("BOM change requested"), "state": b.workflow_state,
			"user": b.modified_by,
		})
	for p in pps:
		history.append({
			"time": p.creation, "doctype": "Production Plan", "name": p.name,
			"event": _("Production Plan updated") if p.production_plan_updated else _("Production Plan linked"),
			"state": p.status, "user": p.modified_by,
		})
	for w in wos:
		history.append({
			"time": w.get("creation"), "doctype": "Work Order", "name": w.get("name"),
			"event": _("Work Order updated") if w.get("modification_status") == "No" else _("Work Order pending update"),
			"state": w.get("status"), "user": w.get("modified_by"),
		})
	history.sort(key=lambda h: str(h["time"]) if h["time"] else "", reverse=True)

	return {
		"so": so,
		"row": row,
		"omrs": omrs,
		"omr_items": omr_items,
		"bmrs": bmrs,
		"pps": pps,
		"wos": wos,
		"pending_actions": pending_actions,
		"history": history,
	}


@frappe.whitelist()
def refresh_row(sales_order, filters=None):
	"""Nothing is cached server side."""
	if isinstance(filters, str):
		try:
			filters = json.loads(filters)
		except Exception:
			filters = {}
	
	filtered_pp = filters.get("production_plan") if filters else None
	maps = _fetch_maps([sales_order], filtered_pp=filtered_pp)
	so = frappe.db.get_value(
		"Sales Order",
		sales_order,
		["name", "customer", "customer_name", "branch", "company", "transaction_date",
		 "delivery_date", "status", "grand_total", "modified", "modified_by"],
		as_dict=True,
	)
	if not so:
		frappe.throw(_("Sales Order {0} not found").format(sales_order))
	return _compute_row(so, maps)


@frappe.whitelist()
def get_charts_data(filters=None):
	rows = _get_computed_rows(_parse_filters(filters), limit=2000)

	# 1) pending updates by branch --------------------------------------
	by_branch = {}
	for r in rows:
		b = r["branch"] or _("Unassigned")
		slot = by_branch.setdefault(b, {"pending": 0, "total": 0})
		slot["total"] += 1
		if r["is_pending"]:
			slot["pending"] += 1
	branch_labels = sorted(by_branch.keys())
	branch_pending = [by_branch[b]["pending"] for b in branch_labels]
	branch_total = [by_branch[b]["total"] for b in branch_labels]

	# 2) status / severity distribution ----------------------------------
	sev_counts = {s: 0 for s in SEVERITY_ORDER}
	for r in rows:
		sev_counts[r["severity"]] += 1

	# 3) manufacturing funnel ---------------------------------------------
	funnel = {
		"SO": len(rows),
		"OMR": sum(1 for r in rows if r["omr_exists"]),
		"BMR": sum(1 for r in rows if r["bmr_exists"]),
		"PP Updated": sum(1 for r in rows if r["pp_updated_flag"]),
		"WO Updated": sum(1 for r in rows if r["wo_synced"]),
		"Completed": sum(1 for r in rows if r["severity"] == "Low" and r["omr_exists"]),
	}

	# 4) 30 day trend -----------------------------------------------------
	trend = {}
	start_date = getdate(add_days(nowdate(), -29))
	d = start_date
	while d <= getdate(nowdate()):
		trend[cstr(d)] = 0
		d = add_days(d, 1)
	for r in rows:
		if not r["is_pending"] or not r["omr"]:
			continue
		created = getdate(r["omr"]["creation"]) if r["omr"].get("creation") else None
		if created and cstr(created) in trend:
			trend[cstr(created)] += 1

	# 5) branch performance --------------------------------------------
	branch_perf = []
	for b in branch_labels:
		total = by_branch[b]["total"]
		pending = by_branch[b]["pending"]
		synced = total - pending
		branch_perf.append({
			"branch": b, "total": total, "synced": synced, "pending": pending,
			"sync_pct": round((synced / total) * 100, 1) if total else 100.0,
		})

	return {
		"branch": {"labels": branch_labels, "pending": branch_pending, "total": branch_total},
		"status_distribution": {"labels": list(sev_counts.keys()), "values": list(sev_counts.values())},
		"funnel": funnel,
		"trend": {"labels": list(trend.keys()), "values": list(trend.values())},
		"branch_performance": branch_perf,
	}


@frappe.whitelist()
def bulk_assign(sales_orders, assign_to, description=None):
	"""Minimal bulk 'Assign' action."""
	names = sales_orders
	if isinstance(names, str):
		names = json.loads(names)

	created = 0
	for so in names:
		if not frappe.db.exists("Sales Order", so):
			continue
		frappe.get_doc({
			"doctype": "ToDo",
			"allocated_to": assign_to,
			"reference_type": "Sales Order",
			"reference_name": so,
			"description": description or _("Please review pending update propagation for {0}").format(so),
		}).insert(ignore_permissions=True)
		created += 1

	return {"created": created}


@frappe.whitelist()
def export_excel(filters=None, sales_orders=None):
	"""Server-side export as a colour-coded .xlsx, matching dashboard badges.
	Only Pending and Critical statuses are highlighted; Updated, Not Required,
	and Waiting are left as plain/default cells.
	"""
	if sales_orders:
		if isinstance(sales_orders, str):
			sales_orders = json.loads(sales_orders)
		so_dicts = frappe.get_all(
			"Sales Order",
			filters={"name": ["in", sales_orders]},
			fields=["name", "customer", "customer_name", "branch", "company",
					"transaction_date", "delivery_date", "status", "grand_total",
					"currency", "owner", "modified", "modified_by"],
		)
		maps = _fetch_maps([s.name for s in so_dicts])
		rows = [_compute_row(s, maps) for s in so_dicts]
	else:
		rows = _get_computed_rows(_parse_filters(filters), limit=5000)

	columns = [
		"Sales Order", "Customer", "OMR", "OMR Status",
		"BMR", "BMR Status", "Production Plan",
		"PP Update Req.", "PP Updated", "Work Order",
		"WO Update Req.", "WO Updated",
		"Pending At", "Severity", "Updated By", "Updated Time", "Remarks",
	]

	wb = Workbook()
	ws = wb.active
	ws.title = "PP & WO Control Report"

	thin = Side(style="thin", color="D9E2EC")
	border = Border(left=thin, right=thin, top=thin, bottom=thin)
	header_fill = PatternFill(start_color="2490EF", end_color="2490EF", fill_type="solid")
	header_font = Font(color="FFFFFF", bold=True, size=11)

	for col_idx, title in enumerate(columns, start=1):
		cell = ws.cell(row=1, column=col_idx, value=title)
		cell.fill = header_fill
		cell.font = header_font
		cell.alignment = Alignment(horizontal="center", vertical="center")
		cell.border = border

	ws.freeze_panes = "A2"
	ws.auto_filter.ref = "A1:{0}1".format(get_column_letter(len(columns)))

	# Columns that get badge-colour highlighting (1-indexed)
	BADGE_COL_MAP = {4: "omr_badge", 6: "bmr_badge", 8: "pp_required_badge",
					  9: "pp_updated_badge", 11: "wo_required_badge", 12: "wo_updated_badge"}
	SEVERITY_COL = 14

	row_idx = 2
	for r in rows:
		bmr_list = [b["name"] for b in r.get("bmr_list", [])]
		wo_list = [w["name"] for w in r.get("wo_list", [])]
		remark = "Awaiting action at {0}".format(r.get("current_stage") or "") if r.get("is_pending") else ""

		values = [
			_clean(r.get("sales_order")),
			_clean(r.get("customer_name")),
			_clean(r["omr"]["name"] if r.get("omr") else ""),
			_clean(r["omr_badge"]["label"] if r.get("omr_badge") else ""),
			_clean(_format_doc_list(bmr_list)),
			_clean(r["bmr_badge"]["label"] if r.get("bmr_badge") else ""),
			_clean(r["pp"]["name"] if r.get("pp") else ""),
			_clean(r["pp_required_badge"]["label"] if r.get("pp_required_badge") else ""),
			_clean(r["pp_updated_badge"]["label"] if r.get("pp_updated_badge") else ""),
			_clean(_format_doc_list(wo_list)),
			_clean(r["wo_required_badge"]["label"] if r.get("wo_required_badge") else ""),
			_clean(r["wo_updated_badge"]["label"] if r.get("wo_updated_badge") else ""),
			_clean(r.get("pending_at")),
			_clean(r.get("severity")),
			_clean(r.get("modified_by")),
			_clean(pretty_date(r.get("modified")) if r.get("modified") else ""),
			_clean(remark),
		]

		for col_idx, val in enumerate(values, start=1):
			cell = ws.cell(row=row_idx, column=col_idx, value=val)
			cell.border = border
			cell.alignment = Alignment(vertical="center")

		colored_cols = set()

		for col_idx, badge_key in BADGE_COL_MAP.items():
			badge = r.get(badge_key)
			if not badge:
				continue
			key = badge.get("key")

			# -----------------------------------------------------------
			# Only color Pending and Critical badges. Updated (green),
			# Not Required (blue), and Waiting (grey) are intentionally
			# left uncolored per request - commented out, not removed,
			# so they can be re-enabled later if needed.
			# -----------------------------------------------------------
			if key not in ("pending", "critical"):
				continue
			# if key in ("updated", "not_required", "waiting"):
			#     continue

			hexcode = BADGE_FILL_HEX[key]
			cell = ws.cell(row=row_idx, column=col_idx)
			cell.fill = PatternFill(start_color=hexcode, end_color=hexcode, fill_type="solid")
			cell.font = Font(color="FFFFFF", bold=True)
			cell.alignment = Alignment(horizontal="center", vertical="center")
			colored_cols.add(col_idx)

		# -----------------------------------------------------------
		# Severity column: only color Critical. Other severities
		# (High/amber, Medium/blue, Waiting/grey, Low/green) are
		# intentionally left uncolored - commented out, not removed.
		# -----------------------------------------------------------
		sev = r.get("severity")
		if sev == "Critical":
			hexcode = SEVERITY_FILL_HEX[sev]
			cell = ws.cell(row=row_idx, column=SEVERITY_COL)
			cell.fill = PatternFill(start_color=hexcode, end_color=hexcode, fill_type="solid")
			cell.font = Font(color="FFFFFF", bold=True)
			cell.alignment = Alignment(horizontal="center", vertical="center")
			colored_cols.add(SEVERITY_COL)
		# elif sev in SEVERITY_FILL_HEX:
		#     hexcode = SEVERITY_FILL_HEX[sev]
		#     cell = ws.cell(row=row_idx, column=SEVERITY_COL)
		#     cell.fill = PatternFill(start_color=hexcode, end_color=hexcode, fill_type="solid")
		#     cell.font = Font(color="1A1300" if sev == "High" else "FFFFFF", bold=True)
		#     cell.alignment = Alignment(horizontal="center", vertical="center")
		#     colored_cols.add(SEVERITY_COL)

		# Light red tint across the rest of a Critical row, echoing the
		# left-border indicator in the UI. Still only triggers for Critical.
		if sev == "Critical":
			tint = PatternFill(start_color="FDEDED", end_color="FDEDED", fill_type="solid")
			for col_idx in range(1, len(columns) + 1):
				if col_idx not in colored_cols:
					ws.cell(row=row_idx, column=col_idx).fill = tint

		row_idx += 1

	widths = [16, 22, 16, 14, 34, 14, 16, 14, 14, 34, 14, 14, 22, 12, 14, 14, 30]
	for i, w in enumerate(widths, start=1):
		ws.column_dimensions[get_column_letter(i)].width = w

	buf = io.BytesIO()
	wb.save(buf)
	buf.seek(0)

	frappe.response["filename"] = "production_plan_wo_control_report.xlsx"
	frappe.response["filecontent"] = buf.getvalue()
	frappe.response["content_type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
	frappe.response["type"] = "download"
	
@frappe.whitelist()
def export_csv(filters=None):
    """Server-side export of the full filtered result set."""
    rows = _get_computed_rows(_parse_filters(filters), limit=5000)



    columns = [
        "Sales Order", "Customer", "OMR", "OMR Status",
        "BMR", "BMR Status", "Production Plan",
        "PP Update Req.", "PP Updated", "Work Order",
        "WO Update Req.", "WO Updated",
        "Pending At", "Severity", "Updated By", "Updated Time", "Remarks",
    ]

    lines = [",".join(columns)]

    for r in rows:
        bmr_list = [b["name"] for b in r.get("bmr_list", [])]
        wo_list = [w["name"] for w in r.get("wo_list", [])]
		 # Match JS drawer/grid logic exactly:
        # r.is_pending ? "Awaiting action at " + r.current_stage : "—"
        if r.get("is_pending"):
            remark = "Awaiting action at {0}".format(r.get("current_stage") or "")
        else:
            remark = ""

        vals = [
            _csv_escape(_clean(r.get("sales_order"))),
            _csv_escape(_clean(r.get("customer_name"))),
            _csv_escape(_clean(r["omr"]["name"] if r.get("omr") else "")),
            _csv_escape(_clean(r["omr_badge"]["label"] if r.get("omr_badge") else "")),
            _csv_escape(_clean(_format_doc_list(bmr_list))),
            _csv_escape(_clean(r["bmr_badge"]["label"] if r.get("bmr_badge") else "")),
            _csv_escape(_clean(r["pp"]["name"] if r.get("pp") else "")),
            _csv_escape(_clean(r["pp_required_badge"]["label"] if r.get("pp_required_badge") else "")),
            _csv_escape(_clean(r["pp_updated_badge"]["label"] if r.get("pp_updated_badge") else "")),
            _csv_escape(_clean(_format_doc_list(wo_list))),
            _csv_escape(_clean(r["wo_required_badge"]["label"] if r.get("wo_required_badge") else "")),
            _csv_escape(_clean(r["wo_updated_badge"]["label"] if r.get("wo_updated_badge") else "")),
            _csv_escape(_clean(r.get("pending_at"))),
            _csv_escape(_clean(r.get("severity"))),
            _csv_escape(_clean(r.get("modified_by"))),
			_csv_escape(_clean(pretty_date(r.get("modified")) if r.get("modified") else "")),  
			_csv_escape(_clean(remark)),
        ]
        lines.append(",".join(vals))

    content = "\n".join(lines)
    frappe.response["doctype"] = "Production Plan WO Control Report"
    frappe.response["filename"] = "production_plan_wo_control_report.csv"
    frappe.response["filecontent"] = content
    frappe.response["type"] = "download"


def _clean(value):
    """Normalise 'empty' placeholder values so exported cells are truly
    blank instead of containing literal dash characters that some CSV
    readers misrender (e.g. '—' showing up as 'â€"')."""
    if value is None:
        return ""
    value = cstr(value).strip()
    if value in ("—", "-", "--", "N/A", "n/a", "None"):
        return ""
    return value


def _format_doc_list(names):
    """Format a list of linked document names for CSV display as a
    single line, pipe-separated, e.g.
    'BOM-SA-AD8S38...-01|BOM-SA-AD8S31...-01|BOM-SA-AD8S22...-01'
    """
    if not names:
        return ""
    return "|".join(names)

def _csv_escape(value):
    """Escape CSV values. Preserves internal newlines as \\r\\n so Excel
    renders them as line breaks within a single cell instead of new rows."""
    if value is None:
        return '""'
    value = cstr(value).replace('"', '""')
    value = value.replace("\r\n", "\n").replace("\n", "\r\n")
    return '"{}"'.format(value)

# --------------------------------------------------------------------------- #
# Internal helpers - ALL FILTERS USE EXACT MATCH (==)
# --------------------------------------------------------------------------- #


def _parse_filters(filters):
	if not filters:
		return {}
	if isinstance(filters, str):
		try:
			return json.loads(filters)
		except Exception:
			return {}
	return filters


def _resolve_period_range(filters):
	"""Turn the client's named `period` into an actual (from_date, to_date)
	pair, computed against the SERVER's current date."""
	period = (filters.get("period") or "").strip().lower()
	if not period or period in ("all", "any", "all_time"):
		return None

	today = getdate(nowdate())

	if period == "today":
		return today, today
	if period == "this_week":
		start = add_days(today, -today.weekday())
		return start, add_days(start, 6)
	if period == "last_week":
		this_week_start = add_days(today, -today.weekday())
		start = add_days(this_week_start, -7)
		return start, add_days(start, 6)
	if period == "this_month":
		return get_first_day(today), get_last_day(today)
	if period == "last_month":
		prev = add_months(today, -1)
		return get_first_day(prev), get_last_day(prev)
	if period == "this_year":
		return getdate(f"{today.year}-01-01"), getdate(f"{today.year}-12-31")
	if period == "last_year":
		y = today.year - 1
		return getdate(f"{y}-01-01"), getdate(f"{y}-12-31")
	if period == "custom":
		f = filters.get("from_date")
		t = filters.get("to_date")
		if f and t:
			return getdate(f), getdate(t)
		return None

	return None


def _resolve_link_filters(filters):
	"""
	Turn omr / bmr / production_plan / work_order filters into a single
	restricted set of Sales Order names (or None = no restriction).
	ALL filters use EXACT MATCH (==), not LIKE.
	"""
	restrict = None

	def _intersect(names):
		nonlocal restrict
		names = set(names)
		restrict = names if restrict is None else (restrict & names)

	# OMR filter - EXACT match
	if filters.get("omr"):
		so = frappe.db.get_value("Order Modification Request", filters["omr"], "sales_order")
		_intersect([so] if so else [])

	# BMR filter - EXACT match
	if filters.get("bmr"):
		parents = frappe.get_all(
			"Sales Order Item For OMR",
			filters={"bom_update_request": filters["bmr"]},  # EXACT match
			pluck="parent",
		)
		sos = []
		if parents:
			sos = frappe.get_all(
				"Order Modification Request", 
				filters={"name": ["in", parents]},  # EXACT match
				pluck="sales_order"
			)
		_intersect(sos)

	# Production Plan filter - EXACT match
	if filters.get("production_plan"):
		sos = frappe.get_all(
			"Production Plan Sales Order", 
			filters={"parent": filters["production_plan"]},  # EXACT match
			pluck="sales_order"
		)
		_intersect(sos)

		# Work Order filter - EXACT match
	if filters.get("work_order"):
		pp = frappe.db.get_value("Work Order", filters["work_order"], "production_plan")
		if pp:
			sos = frappe.get_all(
				"Production Plan Sales Order",
				filters={"parent": pp},
				pluck="sales_order"
			)
			_intersect(sos)
		else:
			so = frappe.db.get_value("Work Order", filters["work_order"], "sales_order")
			if so:
				_intersect([so])
			else:
				_intersect([])

	return restrict


def _get_filtered_sales_orders(filters):
	"""
	Build the base Sales Order query with ALL filters using EXACT MATCH (==).
	"""
	so = frappe.qb.DocType("Sales Order")
	query = (
		frappe.qb.from_(so)
		.select(
			so.name, so.customer, so.customer_name, so.branch, so.company,
			so.transaction_date, so.delivery_date, so.status, so.grand_total,
			so.modified, so.modified_by,
		)
		.where(so.docstatus == 1)
	)

	# ALL link filters use EXACT match (==)
	if filters.get("company"):
		query = query.where(so.company == filters["company"])  # EXACT match
	if filters.get("branch"):
		query = query.where(so.branch == filters["branch"])  # EXACT match
	if filters.get("customer"):
		query = query.where(so.customer == filters["customer"])  # EXACT match
	if filters.get("sales_order"):
		query = query.where(so.name == filters["sales_order"])  # EXACT match

	# Date Period filter
	period_range = _resolve_period_range(filters)
	if period_range:
		query = query.where(so.transaction_date[cstr(period_range[0]):cstr(period_range[1])])

	# Resolve OMR/BMR/PP/WO filters (all exact match internally)
	restrict = _resolve_link_filters(filters)
	if restrict is not None:
		if not restrict:
			return []
		query = query.where(so.name.isin(list(restrict)))

	query = query.orderby(so.modified, order=frappe.qb.desc).limit(2000)
	return query.run(as_dict=True)


def _fetch_maps(so_names, filtered_pp=None):
    """
    Batch-fetch every related doctype for the given Sales Orders.
    
    When filtered_pp is provided, only that specific PP and its WOs are
    included, ensuring the grid respects the Production Plan filter exactly.
    Flow: SO → PP (via Production Plan Sales Order table) → WO (via production_plan field)

    IMPORTANT: a Sales Order can have MORE THAN ONE Production Plan. This
    report displays exactly one Production Plan per Sales Order row (the
    filtered one, or - with no filter applied - the most recently created
    one). Work Orders must be counted for that SAME Production Plan only;
    they must never be summed across every Production Plan the Sales Order
    has ever had, or the count shown will not match the Production Plan
    displayed (this was the root cause of inflated / inconsistent Work
    Order counts, e.g. showing "170" for a row that only shows one PP).
    """
    maps = {
        "omr": {},        # so -> latest OMR dict
        "omr_all": {},    # so -> list of all OMR dicts
        "omr_bmr": {},    # omr name -> list of bmr names
        "bmr": {},        # bmr name -> bmr dict
        "pp": {},         # so -> latest linked PP dict
        "wo": {},         # so -> list of WO dicts (only WOs created from PPs)
    }
    if not so_names:
        return maps

    # Fetch OMRs
    omrs = frappe.get_all(
        "Order Modification Request",
        filters={"sales_order": ["in", so_names]},
        fields=["name", "sales_order", "workflow_state", "docstatus", "creation", "modified", "modified_by"],
        order_by="creation desc",
    )
    for o in omrs:
        maps["omr_all"].setdefault(o.sales_order, []).append(o)
        if o.sales_order not in maps["omr"]:
            maps["omr"][o.sales_order] = o

    # Fetch BMRs through OMR items
    omr_names = [o.name for o in omrs]
    if omr_names:
        items = frappe.get_all(
            "Sales Order Item For OMR",
            filters={"parent": ["in", omr_names], "bom_update_request": ["is", "set"]},
            fields=["parent", "bom_update_request"],
        )
        for it in items:
            if it.bom_update_request:
                maps["omr_bmr"].setdefault(it.parent, [])
                if it.bom_update_request not in maps["omr_bmr"][it.parent]:
                    maps["omr_bmr"][it.parent].append(it.bom_update_request)

        bmr_names = sorted({n for names in maps["omr_bmr"].values() for n in names})
        if bmr_names:
            bmrs = frappe.get_all(
                "Bom Modification Request",
                filters={"name": ["in", bmr_names]},
                fields=["name", "workflow_state", "docstatus", "fg_item_code", "creation", "modified"],
            )
            for b in bmrs:
                maps["bmr"][b.name] = b

    # Fetch Production Plans linked to Sales Orders
    # =====================================================================
    # When a specific PP is filtered, ONLY fetch that PP's link instead of
    # all PPs for the SO. This ensures the correct PP is used in the
    # computed row and only its WOs are fetched below.
    # =====================================================================
    pp_link_filters = {"sales_order": ["in", so_names]}
    if filtered_pp:
        pp_link_filters["parent"] = filtered_pp  # Restrict to the filtered PP only

    pp_links = frappe.get_all(
        "Production Plan Sales Order",
        filters=pp_link_filters,
        fields=["parent", "sales_order", "creation"],
        order_by="creation desc",
    )

    # Collect every PP name touched by these links, so we can fetch their
    # details in one batch query. NOTE: this can still contain MULTIPLE
    # PPs per Sales Order when no filter is applied - that's expected,
    # we just haven't picked the "chosen" one per SO yet.
    pp_names_all = sorted({p.parent for p in pp_links})
    pps_by_name = {}
    if pp_names_all:
        pps = frappe.get_all(
            "Production Plan",
            filters={"name": ["in", pp_names_all]},
            fields=[
                "name", "status", "sales_order_modification", "production_plan_updated",
                "work_order_updated", "bom_modification", "branch", "modified",
            ],
        )
        pps_by_name = {p.name: p for p in pps}

    # Choose exactly ONE Production Plan per Sales Order - the filtered
    # one (only one is possible once pp_link_filters restricts by
    # `parent`), or the most recently created one when no filter is
    # applied (pp_links is ordered by creation desc, so the first link
    # seen per SO is the latest PP for that SO).
    for link in pp_links:
        if link.sales_order not in maps["pp"] and link.parent in pps_by_name:
            maps["pp"][link.sales_order] = pps_by_name[link.parent]

    # =====================================================================
    # FIX: Work Orders are now scoped strictly to each Sales Order's
    # CHOSEN Production Plan (built above), not to every Production Plan
    # ever linked to that Sales Order. Previously, when no PP filter was
    # applied, `pp_names` held every PP for every SO in scope, and a
    # Work Order was attributed to a Sales Order whenever its
    # `sales_order` field matched - regardless of which of that SO's
    # several Production Plans the Work Order actually came from. That
    # silently summed Work Orders across sibling Production Plans into
    # one inflated count, while the row/popup only ever displayed a
    # single Production Plan - hence the mismatch.
    # =====================================================================
    so_by_chosen_pp = {pp["name"]: so for so, pp in maps["pp"].items()}
    chosen_pp_names = list(so_by_chosen_pp.keys())

    if chosen_pp_names:
        wos = frappe.get_all(
            "Work Order",
            filters={
                "production_plan": ["in", chosen_pp_names],  # only the chosen PP(s)
                "docstatus": ["!=", 2],  # Not cancelled
            },
            fields=[
                "name", "sales_order", "production_plan", "status", "modification_status",
                "qty", "produced_qty", "production_item", "item_name", "bom_no",
                "docstatus", "creation", "modified", "modified_by", "planned_start_date",
            ],
        )

        for w in wos:
            # The Sales Order this Work Order's Production Plan is
            # "assigned to" for this report - the single source of truth.
            target_so = so_by_chosen_pp.get(w.production_plan)
            if not target_so or target_so not in so_names:
                continue
            # A Production Plan can aggregate items from more than one
            # Sales Order. Only attribute this Work Order to target_so if
            # the Work Order's own sales_order agrees (or is blank, in
            # which case we trust the PP -> SO link resolved above).
            wo_so = w.sales_order or target_so
            if wo_so == target_so:
                maps["wo"].setdefault(target_so, []).append(w)

    return maps


def _compute_row(so, maps):
	"""Compute all derived status fields for a single Sales Order."""
	so_name = so["name"] if isinstance(so, dict) else so.name
	get = so.get if isinstance(so, dict) else (lambda k, d=None: getattr(so, k, d))

	omr = maps["omr"].get(so_name)
	omr_exists = bool(omr)
	omr_approved = bool(omr and omr.get("workflow_state") == "Approved") if isinstance(omr, dict) else bool(
		omr and omr.workflow_state == "Approved"
	)
	omr_dict = dict(omr) if omr else None

	bmr_names = maps["omr_bmr"].get(omr["name"] if omr_dict else None, []) if omr_dict else []
	bmr_list = [dict(maps["bmr"][n]) for n in bmr_names if n in maps["bmr"]]
	bmr_exists = len(bmr_list) > 0
	bmr_approved_list = [b for b in bmr_list if b.get("workflow_state") == "Approved"]
	bmr_pending_list = [b for b in bmr_list if b.get("workflow_state") != "Approved"]
	bmr_any_approved = len(bmr_approved_list) > 0
	bmr_all_approved = bmr_exists and len(bmr_pending_list) == 0

	pp = maps["pp"].get(so_name)
	pp_dict = dict(pp) if pp else None
	pp_exists = bool(pp_dict)
	pp_updated_flag = bool(pp_dict and cint(pp_dict.get("production_plan_updated")))

	# PP Update Required: BMR approved but PP not updated yet
	pp_update_required = bmr_any_approved and not pp_updated_flag

	# Get WOs linked to this SO (only those created from PPs)
	wo_list = [dict(w) for w in maps["wo"].get(so_name, [])]
	wo_pending_list = [w for w in wo_list if w.get("modification_status") == "Yes"]
	wo_exists = len(wo_list) > 0
	wo_synced = wo_exists and len(wo_pending_list) == 0
	
	# WO Update Required: PP is updated but WOs still have modification_status = "Yes"
	wo_update_required = pp_updated_flag and len(wo_pending_list) > 0

	# ---------------- severity -------------------------------------------
	if pp_update_required:
		severity = "Critical"
	elif wo_update_required:
		severity = "High"
	elif omr_approved and bmr_exists and not bmr_all_approved:
		severity = "Medium"
	elif omr_exists and not omr_approved:
		severity = "Waiting"
	else:
		severity = "Low"

	is_pending = severity in ("Critical", "High", "Medium", "Waiting")

	# ---------------- badges ----------------------------------------------
	if not omr_exists:
		omr_badge = _badge("not_required")
	elif omr_approved:
		omr_badge = _badge("updated")
	else:
		omr_badge = _badge("pending")

	if not bmr_exists:
		bmr_badge = _badge("not_required")
	elif bmr_all_approved:
		bmr_badge = _badge("updated")
	else:
		bmr_badge = _badge("pending")

	if pp_update_required:
		pp_required_badge = _badge("critical")
	elif bmr_any_approved:
		pp_required_badge = _badge("not_required")
	else:
		pp_required_badge = _badge("not_required")

	if not bmr_any_approved:
		pp_updated_badge = _badge("not_required")
	elif pp_updated_flag:
		pp_updated_badge = _badge("updated")
	else:
		pp_updated_badge = _badge("pending")

	if wo_update_required:
		wo_required_badge = _badge("critical") if severity == "High" else _badge("pending")
	else:
		wo_required_badge = _badge("not_required")

	if not pp_updated_flag:
		wo_updated_badge = _badge("not_required")
	elif not wo_exists:
		wo_updated_badge = _badge("waiting")
	elif wo_synced:
		wo_updated_badge = _badge("updated")
	else:
		wo_updated_badge = _badge("pending")

	# ---------------- stage timeline ---------------------------------------
	stages = []
	stages.append({"key": "so", "label": "SO", "state": "completed"})

	if not omr_exists:
		omr_state = "not_required"
	elif omr_approved:
		omr_state = "completed"
	else:
		omr_state = "pending"
	stages.append({"key": "omr", "label": "OMR", "state": omr_state})

	if not bmr_exists:
		bmr_state = "not_required"
	elif bmr_all_approved:
		bmr_state = "completed"
	else:
		bmr_state = "pending"
	stages.append({"key": "bmr", "label": "BMR", "state": bmr_state})

	if not bmr_any_approved:
		pp_state = "not_required"
	elif pp_updated_flag:
		pp_state = "completed"
	else:
		pp_state = "blocked"  # BMR approved but PP not updated - needs "Get Update" click
	stages.append({"key": "pp", "label": "PP", "state": pp_state})

	if not pp_updated_flag:
		wo_state = "not_required"
	elif wo_synced:
		wo_state = "completed"
	else:
		wo_state = "blocked" if wo_exists else "pending"
	stages.append({"key": "wo", "label": "WO", "state": wo_state})

	completed_state = "completed" if severity == "Low" else "pending"
	stages.append({"key": "completed", "label": "Completed", "state": completed_state})

	# ---------------- pending-at / current stage ---------------------------
	pending_at = "—"
	current_stage = "Completed"
	for st in stages:
		if st["key"] in ("so", "completed"):
			continue
		if st["state"] in ("pending", "blocked"):
			pending_at = {
				"omr": _("OMR Approval"),
				"bmr": _("BMR Approval"),
				"pp": _("Production Plan Update (Click 'Get Update')"),
				"wo": _("Work Order Update"),
			}[st["key"]]
			current_stage = STAGE_LABEL[st["key"]]
			break

	return {
		"sales_order": so_name,
		"customer": get("customer"),
		"customer_name": get("customer_name"),
		"branch": get("branch"),
		"company": get("company"),
		"transaction_date": get("transaction_date"),
		"delivery_date": get("delivery_date"),
		"so_status": get("status"),
		"grand_total": flt(get("grand_total")),
		"modified": get("modified"),
		"modified_by": get("modified_by"),

		"omr": omr_dict,
		"omr_exists": omr_exists,
		"omr_approved": omr_approved,
		"omr_badge": omr_badge,

		"bmr_list": bmr_list,
		"bmr_pending": bmr_pending_list,
		"bmr_exists": bmr_exists,
		"bmr_all_approved": bmr_all_approved,
		"bmr_any_approved": bmr_any_approved,
		"bmr_badge": bmr_badge,

		"pp": pp_dict,
		"pp_exists": pp_exists,
		"pp_updated_flag": pp_updated_flag,
		"pp_update_required": pp_update_required,
		"pp_required_badge": pp_required_badge,
		"pp_updated_badge": pp_updated_badge,

		"wo_list": wo_list,
		"wo_exists": wo_exists,
		"wo_synced": wo_synced,
		"wo_update_required": wo_update_required,
		"wo_required_badge": wo_required_badge,
		"wo_updated_badge": wo_updated_badge,

		"severity": severity,
		"severity_color": SEVERITY_COLOR[severity],
		"is_pending": is_pending,
		"stages": stages,
		"pending_at": pending_at,
		"current_stage": current_stage,
	}


def _get_computed_rows(filters, limit=2000):
    """Get all rows with computed status, then apply post-filters."""
    sos = _get_filtered_sales_orders(filters)
    if not sos:
        return []

    so_names = [s.name for s in sos]
    # KEY FIX: Pass the production_plan filter so _fetch_maps restricts
    # to that specific PP and only fetches its WOs
    filtered_pp = filters.get("production_plan")
    maps = _fetch_maps(so_names, filtered_pp=filtered_pp)
    rows = [_compute_row(s, maps) for s in sos]

    # Apply post-computation filters
    if cint(filters.get("pending_only")):
        rows = [r for r in rows if r["is_pending"]]

    if cint(filters.get("critical_only")):
        rows = [r for r in rows if r["severity"] == "Critical"]

    # KPI card filters
    status_filter = filters.get("status")
    if status_filter and status_filter != "Any":
        if status_filter == "Fully Synced":
            rows = [r for r in rows if r["severity"] == "Low"]
        elif status_filter == "Pending Update":
            rows = [r for r in rows if r["is_pending"]]
        elif status_filter == "Waiting on Approval":
            rows = [r for r in rows if r["severity"] == "Waiting"]

    priority_filter = filters.get("priority")
    if priority_filter and priority_filter != "Any":
        rows = [r for r in rows if r["severity"] == priority_filter]

    return rows[:limit]