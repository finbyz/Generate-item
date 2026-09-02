# -*- coding: utf-8 -*-
# Copyright (c) 2026, Steelstrong / Custom App
# License: MIT / see LICENSE
"""
Production Plan & Work Order Update Control Report
====================================================

BATCH-WISE VERSION - HIGH-PERFORMANCE DATA PIPELINE
---------------------------------------------------
The report grain has changed from "one row per Sales Order" to
"one row per changed batch line" (a line inside an Order Modification
Request's item table that actually carries a revision).

Rules implemented (as agreed with the business owner):

1. A "batch" is identified by the `batch_no` field on the
   `Sales Order Item For OMR` child row.
2. A batch line is considered CHANGED if ANY of its `rev_*` fields
   (`rev_item`, `rev_qty`, `rev_description`, `rev_line_status`, and any
   other `rev_*` field present on the row) carries a value that is
   actually DIFFERENT from its corresponding original field. A rev_
   field simply being re-populated with the same value as the original
   does not count as a change. Only changed lines are shown in the
   report. See `_check_item_has_changes()`.
3. If a Sales Order has more than one OMR raised against it over time,
   each OMR is walked independently - so OMR #1 can contribute 5 changed
   batch rows and a later OMR #2 on the very same Sales Order can
   contribute another 3 changed batch rows; both sets show up as
   separate rows in the report, each tagged with its own OMR.
4. BMR, Production Plan and Work Order status shown against a batch row
   are all resolved at BATCH + EFFECTIVE-ITEM level, never merely at
   Sales-Order level:
     - BMR: resolved via the line's `bom_update_request` field.
     - Production Plan: resolved by checking whether the batch's
       effective item actually exists in that Production Plan's
       Assembly Item child table (see `_resolve_pp_for_batch`). A PP
       linked to the same Sales Order is NEVER assumed to apply to a
       batch just because they share the Sales Order.
     - Work Order: resolved by matching `production_item` AND
       `production_plan` on the Work Order against the batch's own
       resolved Production Plan + effective item.
   The "effective item" for a line is `rev_item` when it is genuinely
   different from `item` (an item replacement), otherwise it is the
   original `item` (this covers quantity-only changes too - those
   never require `rev_item` to be populated). See `_effective_item()`.

Flow reminder:
1. OMR submitted -> Production Plan sales_order_modification = "Yes"
2. BMR submitted -> Production Plan bom_modification = "Yes"
3. Production Plan "Get Update" button clicked -> updates applied
4. Work Orders created FROM Production Plans (PP -> WO relationship)
5. Report shows only batch lines that actually changed, one row each
"""

from __future__ import unicode_literals

import json
import io
import hashlib
from datetime import datetime as dt
from datetime import date as date_type

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import frappe
from frappe import _
from frappe.utils import (
	add_days,
	add_months,
	cint,
	cstr,
	flt,
	get_first_day,
	get_last_day,
	getdate,
	nowdate,
	pretty_date,
	now_datetime,
)


# --------------------------------------------------------------------------- #
# Badge / severity vocabulary
# --------------------------------------------------------------------------- #

BADGES = {
	"updated": {"key": "updated", "emoji": "", "label": "Updated", "color": "green"},
	"pending": {"key": "pending", "emoji": "", "label": "Pending", "color": "amber"},
	"critical": {"key": "critical", "emoji": "", "label": "Critical", "color": "red"},
	"not_required": {"key": "not_required", "emoji": "", "label": "Not Required", "color": "blue"},
	"waiting": {"key": "waiting", "emoji": "", "label": "Waiting", "color": "grey"},
}

SEVERITY_ORDER = ["Critical", "High", "Low"]
SEVERITY_COLOR = {
	"Critical": "red",
	"High": "amber",
	"Low": "green",
}
SEVERITY_FILL_HEX = {
	"Critical": "E86161",
	"High": "F5A623",
	"Low": "5EBB63",
}
BADGE_FILL_HEX = {
	"updated": "5EBB63",
	"pending": "F5A623",
	"critical": "E86161",
	"not_required": "4C9CE2",
	"waiting": "98A6BF",
}

STAGE_KEYS = ["so", "omr", "bmr", "pp", "wo", "completed"]
STAGE_LABEL = {
	"so": "SO",
	"omr": "OMR",
	"bmr": "BMR",
	"pp": "PP",
	"wo": "WO",
	"completed": "Completed",
}

PERIOD_OPTIONS = [
	{"value": "", "label": "All Time"},
	{"value": "today", "label": "Today"},
	{"value": "this_week", "label": "This Week"},
	{"value": "last_week", "label": "Last Week"},
	{"value": "this_month", "label": "This Month"},
	{"value": "last_month", "label": "Last Month"},
	{"value": "this_year", "label": "This Year"},
	{"value": "last_year", "label": "Last Year"},
	{"value": "custom", "label": "Custom…"},
]

# --------------------------------------------------------------------------- #
# Assembly Item child-table configuration
# --------------------------------------------------------------------------- #
PP_ASSEMBLY_ITEM_DOCTYPE_OVERRIDE = None

# Process-level cache for detected assembly item doctype
_CACHED_ASSEMBLY_ITEM_DOCTYPE = None

# Cache TTL for _fetch_maps results (seconds) - per SO basis
MAPS_CACHE_TTL = 30

# How long computed report rows are cached (seconds).
ROWS_CACHE_TTL = 30


def _badge(key):
	return dict(BADGES[key])


def _status_badge(status, kind):
	"""Map a computed BMR/PP/WO status string to a badge, mirroring the
	same visual vocabulary that was used in the Sales-Order level report."""
	if kind == "bmr":
		if status in ("Not Created", "Not Required"):
			return _badge("not_required")
		if status in ("Draft", "Pending"):
			return _badge("pending")
		if status == "Rejected":
			return _badge("critical")
		if status == "Updated":
			return _badge("updated")
		return _badge("not_required")
	if kind == "pp":
		if status in ("Not Created", "Not Required"):
			return _badge("not_required")
		if status == "Pending":
			return _badge("pending")
		if status == "Updated":
			return _badge("updated")
		return _badge("not_required")
	if kind == "wo":
		if status in ("Not Created", "Not Started"):
			return _badge("not_required")
		if status == "Pending":
			return _badge("pending")
		if status == "Updated":
			return _badge("updated")
		return _badge("not_required")
	return _badge("not_required")


def _has_value(v):
	"""True if a field carries an actual value (not None / not blank)."""
	if v is None:
		return False
	if isinstance(v, str):
		return v.strip() != ""
	return True


def _norm_val(v):
	"""Normalise a value for comparison: numeric-looking values compare as
	numbers, everything else compares as a stripped string. None/blank -> None.
	Optimized to avoid ValueError exception creation overhead on non-numeric strings."""
	if v is None:
		return None
	if isinstance(v, (int, float)):
		return flt(v)
	s = cstr(v).strip()
	if not s:
		return None
	# Fast-path check for numeric strings before invoking float parsing
	first = s[0]
	if first.isdigit() or (first in "+-." and len(s) > 1):
		try:
			return flt(s)
		except (ValueError, TypeError):
			pass
	return s


def _effective_item(item):
	"""The item code a batch line ACTUALLY resolves to right now, per the
	agreed business rule:

	    effective_item = rev_item if rev_item is genuinely different
	                                 from item, else item
	"""
	item_code = item.get("item")
	rev_item = item.get("rev_item")
	if _has_value(rev_item) and _norm_val(rev_item) != _norm_val(item_code):
		return rev_item
	return item_code


def _check_item_has_changes(item):
	"""A batch line is CHANGED if ANY `rev_*` field on it carries a value
	that is actually different from its corresponding original field
	(rev_item vs item, rev_qty vs qty, rev_description vs description,
	rev_line_status vs line_status, and any other rev_* field present on
	the row). A rev_ field that is merely re-populated with the same value
	as the original does NOT count as a change."""
	for key, rev_val in item.items():
		if not key.startswith("rev_"):
			continue
		if rev_val is None or rev_val == "":
			continue
		base_key = key[4:]  # len("rev_") is 4
		base_val = item.get(base_key)
		if _norm_val(rev_val) != _norm_val(base_val):
			return True
	return False


def _line_change_type(item):
	"""Classify what kind of change this specific line represents."""
	rev_item = item.get("rev_item")
	item_code = item.get("item")
	has_item_replacement = _has_value(item.get("bom_update_request")) or (
		_has_value(rev_item) and _norm_val(rev_item) != _norm_val(item_code)
	)

	rev_qty = item.get("rev_qty")
	qty = item.get("qty")
	has_qty_change = _has_value(rev_qty) and _norm_val(rev_qty) != _norm_val(qty)

	if has_item_replacement:
		return "Item Replacement"
	if has_qty_change:
		return "Quantity Change"
	return "Other"


# --------------------------------------------------------------------------- #
# Public whitelisted API
# --------------------------------------------------------------------------- #


@frappe.whitelist()
def get_filter_options():
	"""Static option lists for the sticky filter bar."""
	companies = frappe.get_all("Company", pluck="name", order_by="name")
	if frappe.db.table_exists("Branch"):
		branches = frappe.get_all("Branch", pluck="name", order_by="name")
	else:
		branches = frappe.get_all(
			"Sales Order", pluck="branch", distinct=True, filters={"branch": ["is", "set"]}
		)

	return {
		"companies": companies,
		"branches": [b for b in branches if b],
		"period_options": PERIOD_OPTIONS,
		"default_from": add_days(nowdate(), -90),
		"default_to": nowdate(),
	}


def _compute_dashboard_summary(rows):
	"""Compute KPI cards + overall sync ring over computed batch rows in a single pass."""
	pending_pp = 0
	pending_wo = 0
	pending_bmr = 0
	pending_omr = 0
	completed = 0
	critical = 0

	for r in rows:
		if r["pp_status"] in ("Pending", "Not Created"):
			pending_pp += 1
		if r["wo_status"] in ("Pending", "Not Created", "Not Started"):
			pending_wo += 1
		if r["bmr_status"] in ("Draft", "Pending", "Not Created") and r["change_type"] == "Item Replacement":
			pending_bmr += 1
		if r["omr_workflow_state"] != "Approved":
			pending_omr += 1
		if not r["is_pending"]:
			completed += 1
		if r["severity"] == "Critical":
			critical += 1

	trackable = len(rows)
	sync_pct = round((completed / trackable) * 100, 1) if trackable else 100.0

	return {
		"pending_pp": pending_pp,
		"pending_wo": pending_wo,
		"pending_bmr": pending_bmr,
		"pending_omr": pending_omr,
		"completed": completed,
		"critical": critical,
		"total": len(rows),
		"trackable": trackable,
		"sync_pct": sync_pct,
		"all_clear": (
			trackable > 0
			and pending_pp == 0
			and pending_wo == 0
			and pending_bmr == 0
			and pending_omr == 0
			and critical == 0
		),
	}


def _compute_charts_summary(rows):
	"""Compute all chart datasets in a single pass over rows."""
	by_branch = {}
	sev_counts = {s: 0 for s in SEVERITY_ORDER}
	omr_approved = 0
	bmr_updated = 0
	pp_updated = 0
	wo_updated = 0
	completed = 0

	# 30-day trend date buckets
	today_dt = getdate(nowdate())
	start_dt = getdate(add_days(nowdate(), -29))
	trend = {}
	d = start_dt
	while d <= today_dt:
		trend[cstr(d)] = 0
		d = add_days(d, 1)

	for r in rows:
		# Branch counts
		b = r.get("branch") or _("Unassigned")
		slot = by_branch.setdefault(b, {"pending": 0, "total": 0})
		slot["total"] += 1
		if r["is_pending"]:
			slot["pending"] += 1

		# Severity distribution
		sev = r.get("severity")
		if sev in sev_counts:
			sev_counts[sev] += 1
		else:
			sev_counts[sev] = 1

		# Manufacturing Funnel
		if r["omr_workflow_state"] == "Approved":
			omr_approved += 1
		if r["bmr_status"] == "Updated" or r["change_type"] != "Item Replacement":
			bmr_updated += 1
		if r["pp_status"] == "Updated":
			pp_updated += 1
		if r["wo_status"] == "Updated":
			wo_updated += 1
		if not r["is_pending"]:
			completed += 1

		# 30-day pending trend
		if r["is_pending"] and r.get("omr_creation"):
			created = getdate(r["omr_creation"])
			created_str = cstr(created)
			if created_str in trend:
				trend[created_str] += 1

	branch_labels = sorted(by_branch.keys())
	branch_pending = [by_branch[b]["pending"] for b in branch_labels]
	branch_total = [by_branch[b]["total"] for b in branch_labels]

	funnel = {
		"Batch Lines": len(rows),
		"OMR Approved": omr_approved,
		"BMR Updated": bmr_updated,
		"PP Updated": pp_updated,
		"WO Updated": wo_updated,
		"Completed": completed,
	}

	branch_perf = []
	for b in branch_labels:
		tot = by_branch[b]["total"]
		pen = by_branch[b]["pending"]
		syn = tot - pen
		branch_perf.append({
			"branch": b,
			"total": tot,
			"synced": syn,
			"pending": pen,
			"sync_pct": round((syn / tot) * 100, 1) if tot else 100.0,
		})

	return {
		"branch": {"labels": branch_labels, "pending": branch_pending, "total": branch_total},
		"status_distribution": {"labels": list(sev_counts.keys()), "values": list(sev_counts.values())},
		"funnel": funnel,
		"trend": {"labels": list(trend.keys()), "values": list(trend.values())},
		"branch_performance": branch_perf,
	}


@frappe.whitelist()
def get_all_data(filters=None, start=0, page_length=50):
	"""High-performance unified endpoint to fetch Dashboard KPIs, Paginated Grid,
	and Charts in a SINGLE roundtrip, completely eliminating parallel query stampedes."""
	parsed_filters = _parse_filters(filters)
	rows = _get_computed_rows(parsed_filters, limit=5000)

	start = cint(start)
	page_length = cint(page_length) or 50

	total = len(rows)
	page = rows[start : start + page_length]

	return {
		"dashboard": _compute_dashboard_summary(rows),
		"grid": {"rows": page, "total": total, "start": start, "page_length": page_length},
		"charts": _compute_charts_summary(rows),
	}


@frappe.whitelist()
def get_dashboard_data(filters=None):
	"""KPI cards + overall sync ring - now computed over batch rows."""
	rows = _get_computed_rows(_parse_filters(filters), limit=5000)
	return _compute_dashboard_summary(rows)


@frappe.whitelist()
def get_grid_data(filters=None, start=0, page_length=50):
	"""Paginated grid rows (batch-line grain)."""
	start = cint(start)
	page_length = cint(page_length) or 50

	rows = _get_computed_rows(_parse_filters(filters), limit=5000)
	total = len(rows)
	page = rows[start : start + page_length]

	return {"rows": page, "total": total, "start": start, "page_length": page_length}


@frappe.whitelist()
def get_charts_data(filters=None):
	"""Charts dataset calculation."""
	rows = _get_computed_rows(_parse_filters(filters), limit=5000)
	return _compute_charts_summary(rows)


@frappe.whitelist()
def get_row_detail(sales_order, filters=None, batch_no=None, item=None):
	"""Everything the right-hand drawer needs for a specific batch row.
	
	If batch_no and item are provided, only shows data for that specific
	batch+item combination. Otherwise shows all changed batch rows for the SO.
	"""
	if not sales_order:
		frappe.throw(_("Sales Order is required"))

	if isinstance(filters, str):
		try:
			filters = json.loads(filters)
		except Exception:
			filters = {}

	so = frappe.db.get_value(
		"Sales Order",
		sales_order,
		[
			"name", "customer", "customer_name", "branch", "company", "transaction_date",
			"delivery_date", "status", "grand_total", "currency", "owner", "modified", "modified_by",
		],
		as_dict=True,
	)
	if not so:
		frappe.throw(_("Sales Order {0} not found").format(sales_order))

	filtered_pp = filters.get("production_plan") if filters else None

	# Always initialise maps first so it is always defined regardless of cache path
	maps = _fetch_maps([sales_order], filtered_pp=filtered_pp)

	# Try to get rows from cache first (computed by grid/dashboard/charts)
	raw_rows = _get_raw_computed_rows(filters or {}, limit=10000)
	batch_rows = [r for r in raw_rows if r["sales_order"] == sales_order]

	# Filter to specific batch+item if provided (for row-click drawer)
	if batch_no and item:
		batch_rows = [
			r for r in batch_rows 
			if (r.get("batch_no") or "") == (batch_no or "")
			and (r.get("item") or "") == (item or "")
		]

	# If not in cache (different filters or cache miss), fall back to fresh compute
	if not batch_rows:
		batch_rows = _compute_batch_rows_for_so(so, maps)
		if batch_no and item:
			batch_rows = [
				r for r in batch_rows 
				if (r.get("batch_no") or "") == (batch_no or "")
				and (r.get("item") or "") == (item or "")
			]

	omrs = maps.get("omr_all", {}).get(sales_order, [])
	
	# Filter OMRs to only those that have the specific batch+item
	if batch_no and item:
		omr_names_with_batch = {r["omr"] for r in batch_rows}
		omrs = [o for o in omrs if o.get("name") in omr_names_with_batch]

	# Get BMRs from maps (fetched in _fetch_maps)
	bmr_names_seen = sorted({r["bmr"]["name"] for r in batch_rows if r.get("bmr")})
	bmrs = [maps.get("bmr", {}).get(n) for n in bmr_names_seen if n in maps.get("bmr", {})]

	pps_seen = {}
	for r in batch_rows:
		if r.get("pp"):
			pps_seen[r["pp"]["name"]] = r["pp"]
	pps = list(pps_seen.values())

	wos_seen = {}
	for r in batch_rows:
		for w in (r.get("wo_list") or []):
			wos_seen[w["name"]] = w
	wos = list(wos_seen.values())

	history = []
	for o in omrs:
		history.append({
			"time": o.get("creation"), "doctype": "Order Modification Request", "name": o.get("name"),
			"event": _("Change request raised"), "state": o.get("workflow_state"),
			"user": o.get("owner"),
		})
	for b in bmrs:
		if b:
			history.append({
				"time": b.get("creation"), "doctype": "Bom Modification Request", "name": b.get("name"),
				"event": _("BOM change requested"), "state": b.get("workflow_state"),
				"user": b.get("modified_by"),
			})
	for p in pps:
		history.append({
			"time": p.get("creation"), "doctype": "Production Plan", "name": p.get("name"),
			"event": _("Production Plan updated") if p.get("production_plan_updated") else _("Production Plan linked"),
			"state": p.get("status"), "user": p.get("modified_by"),
		})
	for w in wos:
		history.append({
			"time": w.get("creation"), "doctype": "Work Order", "name": w.get("name"),
			"event": _("Work Order updated") if w.get("modification_status") == "No" else _("Work Order pending update"),
			"state": w.get("status"), "user": w.get("modified_by"),
		})
	history.sort(key=lambda h: str(h["time"]) if h["time"] else "", reverse=True)

	return {
		"so": so,
		"batch_rows": batch_rows,
		"omrs": omrs,
		"bmrs": bmrs,
		"pps": pps,
		"wos": wos,
		"history": history,
	}


@frappe.whitelist()
def refresh_row(sales_order, filters=None):
	"""Recompute every changed batch row for a Sales Order.
	Busts all relevant caches so the next fetch picks up the change immediately."""
	if isinstance(filters, str):
		try:
			filters = json.loads(filters)
		except Exception:
			filters = {}

	filtered_pp = filters.get("production_plan") if filters else None

	# Bust the maps cache for this SO
	_bust_maps_cache([sales_order], filtered_pp)

	# Bust the rows cache (used by grid/dashboard/charts)
	_bust_rows_cache()

	# Recompute fresh
	maps = _fetch_maps([sales_order], filtered_pp=filtered_pp)
	so = frappe.db.get_value(
		"Sales Order",
		sales_order,
		["name", "customer", "customer_name", "branch", "company", "transaction_date",
		 "delivery_date", "status", "grand_total", "modified", "modified_by"],
		as_dict=True,
	)
	if not so:
		frappe.throw(_("Sales Order {0} not found").format(sales_order))

	return {"rows": _compute_batch_rows_for_so(so, maps)}


def _bust_maps_cache(so_names, filtered_pp=None):
	"""Invalidate cached _fetch_maps results for the given SOs."""
	try:
		cache_key = "ppwo_maps:{0}:{1}".format(
			hashlib.md5(json.dumps(sorted(so_names), sort_keys=True).encode()).hexdigest()[:16],
			filtered_pp or "none"
		)
		frappe.cache().delete_value(cache_key)
	except Exception:
		pass


@frappe.whitelist()
def bulk_assign(row_ids, assign_to, description=None):
	"""Minimal bulk 'Assign' action - now keyed by batch row id (the OMR
	item child-row name), deduplicated by Sales Order for the ToDo."""
	ids = row_ids
	if isinstance(ids, str):
		ids = json.loads(ids)

	sales_orders = sorted({rid.split("::")[0] for rid in ids if "::" in rid})

	created = 0
	for so in sales_orders:
		if not frappe.db.exists("Sales Order", so):
			continue
		frappe.get_doc({
			"doctype": "ToDo",
			"allocated_to": assign_to,
			"reference_type": "Sales Order",
			"reference_name": so,
			"description": description or _("Please review pending update propagation for {0}").format(so),
		}).insert(ignore_permissions=True)
		created += 1

	return {"created": created}


@frappe.whitelist()
def export_excel(filters=None, row_ids=None):
	"""Server-side export as a colour-coded .xlsx, batch-row grain."""
	if row_ids:
		if isinstance(row_ids, str):
			row_ids = json.loads(row_ids)
		row_id_set = set(row_ids)
		rows = [r for r in _get_computed_rows(_parse_filters(filters), limit=5000) if r["row_id"] in row_id_set]
	else:
		rows = _get_computed_rows(_parse_filters(filters), limit=10000)

	columns = [
		"Sales Order", "Customer", "OMR", "Batch No", "Item",
		"BMR", "BMR Status", "Production Plan", "PP Status",
		"Work Order", "WO Status", "Severity", "Remarks", "Updated By", "Updated Time",
	]

	wb = Workbook()
	ws = wb.active
	ws.title = "PP & WO Control Report"

	thin = Side(style="thin", color="D9E2EC")
	border = Border(left=thin, right=thin, top=thin, bottom=thin)
	header_fill = PatternFill(start_color="2490EF", end_color="2490EF", fill_type="solid")
	header_font = Font(color="FFFFFF", bold=True, size=11)

	for col_idx, title in enumerate(columns, start=1):
		cell = ws.cell(row=1, column=col_idx, value=title)
		cell.fill = header_fill
		cell.font = header_font
		cell.alignment = Alignment(horizontal="center", vertical="center")
		cell.border = border

	ws.freeze_panes = "A2"
	ws.auto_filter.ref = "A1:{0}1".format(get_column_letter(len(columns)))

	BADGE_COL_MAP = {7: "bmr_status_badge", 9: "pp_status_badge", 11: "wo_status_badge"}
	SEVERITY_COL = 12

	row_idx = 2
	for r in rows:
		wo_list = [w["name"] for w in r.get("wo_list", [])]

		values = [
			_clean(r.get("sales_order")),
			_clean(r.get("customer_name")),
			_clean(r.get("omr")),
			_clean(r.get("batch_no")),
			_clean(r.get("item")),
			_clean(r["bmr"]["name"] if r.get("bmr") else ""),
			_clean(r.get("bmr_status")),
			_clean(r["pp"]["name"] if r.get("pp") else ""),
			_clean(r.get("pp_status")),
			_clean(_format_doc_list(wo_list)),
			_clean(r.get("wo_status")),
			_clean(r.get("severity")),
			_clean(r.get("remarks")),
			_clean(r.get("modified_by")),
			_clean(pretty_date(r.get("modified")) if r.get("modified") else ""),
		]

		for col_idx, val in enumerate(values, start=1):
			cell = ws.cell(row=row_idx, column=col_idx, value=val)
			cell.border = border
			cell.alignment = Alignment(vertical="center")

		colored_cols = set()

		for col_idx, badge_key in BADGE_COL_MAP.items():
			badge = r.get(badge_key)
			if not badge:
				continue
			key = badge.get("key")
			if key not in ("pending", "critical"):
				continue

			hexcode = BADGE_FILL_HEX[key]
			cell = ws.cell(row=row_idx, column=col_idx)
			cell.fill = PatternFill(start_color=hexcode, end_color=hexcode, fill_type="solid")
			cell.font = Font(color="FFFFFF", bold=True)
			cell.alignment = Alignment(horizontal="center", vertical="center")
			colored_cols.add(col_idx)

		sev = r.get("severity")
		if sev == "Critical":
			hexcode = SEVERITY_FILL_HEX[sev]
			cell = ws.cell(row=row_idx, column=SEVERITY_COL)
			cell.fill = PatternFill(start_color=hexcode, end_color=hexcode, fill_type="solid")
			cell.font = Font(color="FFFFFF", bold=True)
			cell.alignment = Alignment(horizontal="center", vertical="center")
			colored_cols.add(SEVERITY_COL)

		if sev == "Critical":
			tint = PatternFill(start_color="FDEDED", end_color="FDEDED", fill_type="solid")
			for col_idx in range(1, len(columns) + 1):
				if col_idx not in colored_cols:
					ws.cell(row=row_idx, column=col_idx).fill = tint

		row_idx += 1

	widths = [16, 22, 16, 14, 18, 18, 10, 10, 16, 14, 16, 14, 18, 14, 12, 34, 14, 14]
	for i, w in enumerate(widths, start=1):
		ws.column_dimensions[get_column_letter(i)].width = w

	buf = io.BytesIO()
	wb.save(buf)
	buf.seek(0)

	frappe.response["filename"] = "production_plan_wo_control_report.xlsx"
	frappe.response["filecontent"] = buf.getvalue()
	frappe.response["content_type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
	frappe.response["type"] = "download"


def _clean(value):
	"""Normalise 'empty' placeholder values."""
	if value is None:
		return ""
	value = cstr(value).strip()
	if value in ("—", "-", "--", "N/A", "n/a", "None"):
		return ""
	return value


def _format_doc_list(names):
	"""Format a list of linked document names for display."""
	if not names:
		return ""
	return "|".join(names)


# --------------------------------------------------------------------------- #
# Internal helpers
# --------------------------------------------------------------------------- #


def _parse_filters(filters):
	if not filters:
		return {}
	if isinstance(filters, str):
		try:
			return json.loads(filters)
		except Exception:
			return {}
	return filters


def _resolve_period_range(filters):
	"""Turn the client's named `period` into an actual (from_date, to_date) pair."""
	period = (filters.get("period") or "").strip().lower()
	if not period or period in ("all", "any", "all_time"):
		return None

	today = getdate(nowdate())

	if period == "today":
		return today, today
	if period == "this_week":
		start = add_days(today, -today.weekday())
		return start, add_days(start, 6)
	if period == "last_week":
		this_week_start = add_days(today, -today.weekday())
		start = add_days(this_week_start, -7)
		return start, add_days(start, 6)
	if period == "this_month":
		return get_first_day(today), get_last_day(today)
	if period == "last_month":
		prev = add_months(today, -1)
		return get_first_day(prev), get_last_day(prev)
	if period == "this_year":
		return getdate(f"{today.year}-01-01"), getdate(f"{today.year}-12-31")
	if period == "last_year":
		y = today.year - 1
		return getdate(f"{y}-01-01"), getdate(f"{y}-12-31")
	if period == "custom":
		f = filters.get("from_date")
		t = filters.get("to_date")
		if f and t:
			return getdate(f), getdate(t)
		return None

	return None


def _resolve_link_filters(filters):
	"""Turn link filters into a restricted set of Sales Order names."""
	restrict = None

	def _intersect(names):
		nonlocal restrict
		names = set(names)
		restrict = names if restrict is None else (restrict & names)

	if filters.get("omr"):
		so = frappe.db.get_value("Order Modification Request", filters["omr"], "sales_order")
		_intersect([so] if so else [])

	if filters.get("bmr"):
		parents = frappe.get_all(
			"Sales Order Item For OMR",
			filters={"bom_update_request": filters["bmr"]},
			pluck="parent",
		)
		sos = []
		if parents:
			sos = frappe.get_all(
				"Order Modification Request",
				filters={"name": ["in", parents]},
				pluck="sales_order"
			)
		_intersect(sos)

	if filters.get("production_plan"):
		sos = frappe.get_all(
			"Production Plan Sales Order",
			filters={"parent": filters["production_plan"]},
			pluck="sales_order"
		)
		_intersect(sos)

	if filters.get("work_order"):
		pp = frappe.db.get_value("Work Order", filters["work_order"], "production_plan")
		if pp:
			sos = frappe.get_all(
				"Production Plan Sales Order",
				filters={"parent": pp},
				pluck="sales_order"
			)
			_intersect(sos)
		else:
			so = frappe.db.get_value("Work Order", filters["work_order"], "sales_order")
			if so:
				_intersect([so])
			else:
				_intersect([])

	return restrict


def _chunked_get_all(doctype, filters_field, values, fields, order_by=None, additional_filters=None):
	"""Safely query in chunks of 1000 IDs to avoid query packet / parameter limits on huge datasets."""
	if not values:
		return []
	values_list = list(values)
	chunk_size = 1000
	all_results = []
	for i in range(0, len(values_list), chunk_size):
		chunk = values_list[i : i + chunk_size]
		f = {filters_field: ["in", chunk]}
		if additional_filters:
			f.update(additional_filters)
		kwargs = {"filters": f, "fields": fields}
		if order_by:
			kwargs["order_by"] = order_by
		res = frappe.get_all(doctype, **kwargs)
		all_results.extend(res)
	return all_results


def _get_filtered_sales_orders(filters):
	"""Build the base Sales Order query with all filters.
	HIGH-PERFORMANCE OPTIMIZATION: Only fetch Sales Orders that actually
	have submitted Order Modification Requests (docstatus=1), because ONLY
	those can contribute rows to this report."""
	restrict = _resolve_link_filters(filters)
	if restrict is not None and not restrict:
		return []

	omr = frappe.qb.DocType("Order Modification Request")
	so = frappe.qb.DocType("Sales Order")

	query = (
		frappe.qb.from_(so)
		.join(omr)
		.on((omr.sales_order == so.name) & (omr.docstatus == 1))
		.select(
			so.name, so.customer, so.customer_name, so.branch, so.company,
			so.transaction_date, so.delivery_date, so.status, so.grand_total,
			so.modified, so.modified_by,
		)
		.distinct()
		.where(so.docstatus == 1)
	)

	if filters.get("company"):
		query = query.where(so.company == filters["company"])
	if filters.get("branch"):
		query = query.where(so.branch == filters["branch"])
	if filters.get("customer"):
		query = query.where(so.customer == filters["customer"])
	if filters.get("sales_order"):
		query = query.where(so.name == filters["sales_order"])

	period_range = _resolve_period_range(filters)
	if period_range:
		query = query.where(so.transaction_date[cstr(period_range[0]):cstr(period_range[1])])

	if restrict is not None:
		query = query.where(so.name.isin(list(restrict)))

	query = query.orderby(so.modified, order=frappe.qb.desc).limit(5000)
	result = query.run(as_dict=True)
	return result


def _get_pp_assembly_item_doctype():
	"""Auto-detect the Production Plan child table that holds Assembly Item rows.
	Cached at module level and request level to eliminate reflection overhead."""
	global _CACHED_ASSEMBLY_ITEM_DOCTYPE
	if PP_ASSEMBLY_ITEM_DOCTYPE_OVERRIDE:
		return PP_ASSEMBLY_ITEM_DOCTYPE_OVERRIDE

	if _CACHED_ASSEMBLY_ITEM_DOCTYPE:
		return _CACHED_ASSEMBLY_ITEM_DOCTYPE

	cached = getattr(frappe.local, "_ppwo_assembly_item_doctype", "unset")
	if cached != "unset":
		return cached

	candidate = None
	detected_tables = []
	try:
		meta = frappe.get_meta("Production Plan")
		for df in meta.get_table_fields():
			try:
				child_meta = frappe.get_meta(df.options)
			except Exception:
				continue
			item_field = None
			for field_name in ["item_code", "item", "production_item"]:
				if child_meta.has_field(field_name):
					item_field = field_name
					break
			if not item_field:
				continue
			detected_tables.append({
				"doctype": df.options,
				"fieldname": df.fieldname,
				"label": df.label,
				"item_field": item_field,
				"has_rev_item": child_meta.has_field("rev_item"),
				"has_batch_no": child_meta.has_field("batch_no") or child_meta.has_field("custom_batch_no"),
			})
			haystack = "{0} {1}".format(df.fieldname or "", df.label or "").lower()
			if "assembly" in haystack:
				candidate = df.options
				break
			if not candidate:
				candidate = df.options
	except Exception as e:
		frappe.log_error(
			title="PPWO Control Report: Assembly Item detection error",
			message=frappe.get_traceback(),
		)
		candidate = None

	_CACHED_ASSEMBLY_ITEM_DOCTYPE = candidate
	frappe.local._ppwo_assembly_item_doctype = candidate
	return candidate


def _fetch_maps(so_names, filtered_pp=None):
	"""Batch-fetch every related doctype for the given Sales Orders.

	HIGH-PERFORMANCE DATA INDEXING:
	- Pre-indexes `maps["wo_by_pp_item"][(pp_name, production_item)] = [w, ...]` for instant O(1) matching.
	- Stores `maps["pp_assembly_items"][pp_name]` directly as Python sets for O(1) batch item resolution.
	- Uses chunked querying for resilience with large ID lists.
	"""
	if not so_names:
		return {
			"omr": {},
			"omr_all": {},
			"omr_bmr": {},
			"omr_items": {},
			"bmr": {},
			"pp_all": {},
			"pp_assembly_items": {},
			"wo_all": {},
			"wo_by_pp_item": {},
			"assembly_doctype_found": False,
		}

	cache_key = "ppwo_maps:{0}:{1}".format(
		hashlib.md5(json.dumps(sorted(so_names), sort_keys=True).encode()).hexdigest()[:16],
		filtered_pp or "none"
	)

	cached = frappe.cache().get_value(cache_key)
	if cached is not None:
		# Convert cached lists back to sets for O(1) lookups
		if isinstance(cached.get("pp_assembly_items"), dict):
			cached["pp_assembly_items"] = {
				k: set(v) if isinstance(v, (list, set)) else set()
				for k, v in cached["pp_assembly_items"].items()
			}
		# Rebuild wo_by_pp_item if not in cache
		if "wo_by_pp_item" not in cached:
			cached["wo_by_pp_item"] = {}
			for w_list in cached.get("wo_all", {}).values():
				for w in w_list:
					k = (w.get("production_plan"), w.get("production_item"))
					cached["wo_by_pp_item"].setdefault(k, []).append(w)
		return cached

	maps = {
		"omr": {},
		"omr_all": {},
		"omr_bmr": {},
		"omr_items": {},
		"bmr": {},
		"pp_all": {},
		"pp_assembly_items": {},
		"wo_all": {},
		"wo_by_pp_item": {},
		"assembly_doctype_found": False,
	}

	so_names_set = set(so_names)

	omrs = _chunked_get_all(
		"Order Modification Request",
		"sales_order",
		so_names,
		fields=[
			"name", "sales_order", "workflow_state", "docstatus",
			"modification_type", "type", "creation", "owner", "modified", "modified_by"
		],
		order_by="creation desc",
		additional_filters={"docstatus": 1}
	)

	for o in omrs:
		maps["omr_all"].setdefault(o.sales_order, []).append(o)
		if o.sales_order not in maps["omr"]:
			maps["omr"][o.sales_order] = o

	omr_names = [o.name for o in omrs]
	if omr_names:
		items = _chunked_get_all(
			"Sales Order Item For OMR",
			"parent",
			omr_names,
			fields=[
				"name", "parent", "item", "rev_item", "description", "rev_description",
				"qty", "rev_qty", "bom_update_request", "tag_no", "batch_no",
				"line_status", "rev_line_status", "parenttype", "parentfield"
			],
		)

		for item in items:
			maps["omr_items"].setdefault(item.parent, []).append(item)
			if item.get("bom_update_request"):
				maps["omr_bmr"].setdefault(item.parent, [])
				if item.bom_update_request not in maps["omr_bmr"][item.parent]:
					maps["omr_bmr"][item.parent].append(item.bom_update_request)

		bmr_names = sorted({n for names in maps["omr_bmr"].values() for n in names})
		if bmr_names:
			bmrs = _chunked_get_all(
				"Bom Modification Request",
				"name",
				bmr_names,
				fields=[
					"name", "workflow_state", "docstatus", "fg_item_code",
					"fg_item_name", "item_description", "reason_for_change",
					"batch_no_ref", "creation", "modified", "modified_by"
				],
			)
			for b in bmrs:
				maps["bmr"][b.name] = b

	# ---- ALL Production Plans linked to each Sales Order ----
	pp_link_additional = {"parent": filtered_pp} if filtered_pp else None
	pp_links = _chunked_get_all(
		"Production Plan Sales Order",
		"sales_order",
		so_names,
		fields=["parent", "sales_order", "creation"],
		order_by="creation desc",
		additional_filters=pp_link_additional
	)

	pp_names_all = sorted({p.parent for p in pp_links})
	pps_by_name = {}
	if pp_names_all:
		pps = _chunked_get_all(
			"Production Plan",
			"name",
			pp_names_all,
			fields=[
				"name", "status", "sales_order_modification", "production_plan_updated",
				"work_order_updated", "bom_modification", "branch", "modified",
				"modified_by", "creation", "docstatus"
			],
		)
		pps_by_name = {p.name: p for p in pps}

	for link in pp_links:
		if link.parent in pps_by_name:
			maps["pp_all"].setdefault(link.sales_order, []).append(pps_by_name[link.parent])

	pp_to_sos = {}
	for link in pp_links:
		pp_to_sos.setdefault(link.parent, []).append(link.sales_order)

	# ---- Assembly Item child table ----
	assembly_doctype = _get_pp_assembly_item_doctype()
	maps["assembly_doctype_found"] = bool(assembly_doctype)

	if pp_names_all and assembly_doctype:
		child_meta = frappe.get_meta(assembly_doctype)
		child_fields = ["parent"]

		item_field = None
		for fname in ["item_code", "item", "production_item"]:
			if child_meta.has_field(fname):
				item_field = fname
				child_fields.append(fname)
				break

		batch_field = None
		for fname in ["batch_no", "custom_batch_no"]:
			if child_meta.has_field(fname):
				batch_field = fname
				child_fields.append(fname)
				break

		if child_meta.has_field("rev_item"):
			child_fields.append("rev_item")

		assembly_rows = _chunked_get_all(
			assembly_doctype,
			"parent",
			pp_names_all,
			fields=child_fields,
		)
		for row in assembly_rows:
			items_in_pp = maps["pp_assembly_items"].setdefault(row.parent, set())
			item_code = row.get(item_field) if item_field else None
			batch_no = row.get(batch_field) if batch_field else None
			if item_code:
				items_in_pp.add("{0}|{1}".format(item_code, batch_no or ""))
				items_in_pp.add(item_code)
			if row.get("rev_item"):
				items_in_pp.add("{0}|{1}".format(row.get("rev_item"), batch_no or ""))
				items_in_pp.add(row.get("rev_item"))

	# ---- Work Orders linked to ANY of the Production Plans ----
	if pp_names_all:
		wos = _chunked_get_all(
			"Work Order",
			"production_plan",
			pp_names_all,
			fields=[
				"name", "sales_order", "production_plan", "status",
				"modification_status",
				"qty", "produced_qty", "production_item", "item_name", "bom_no",
				"docstatus", "creation", "modified", "modified_by", "planned_start_date"
			],
			additional_filters={"docstatus": ["!=", 2]}
		)

		for w in wos:
			for target_so in pp_to_sos.get(w.production_plan, []):
				if target_so in so_names_set:
					maps["wo_all"].setdefault(target_so, []).append(w)
			# Pre-index for O(1) match in _compute_batch_row
			k = (w.get("production_plan"), w.get("production_item"))
			maps["wo_by_pp_item"].setdefault(k, []).append(w)

	# Cache payload with serializable lists for redis
	cache_payload = dict(maps)
	cache_payload["pp_assembly_items"] = {k: list(v) for k, v in maps["pp_assembly_items"].items()}
	# Exclude tuple keys from redis cache JSON serialization
	if "wo_by_pp_item" in cache_payload:
		del cache_payload["wo_by_pp_item"]

	frappe.cache().set_value(cache_key, cache_payload, expires_in_sec=MAPS_CACHE_TTL)
	return maps


def _resolve_pp_for_batch(so_name, effective_item, item_batch_no, maps):
	"""Find the Production Plan (if any) that actually applies to this specific batch row."""
	if not effective_item:
		return None

	candidates = maps.get("pp_all", {}).get(so_name, [])
	if not candidates:
		return None

	if maps.get("assembly_doctype_found"):
		pp_assembly_items = maps.get("pp_assembly_items", {})
		batch_key = "{0}|{1}".format(effective_item, item_batch_no or "")

		for pp in candidates:
			items_in_pp = pp_assembly_items.get(pp["name"])
			if not items_in_pp:
				continue
			# Fast set membership check
			if batch_key in items_in_pp or effective_item in items_in_pp:
				return pp
		return None

	return None


def _pp_update_flags(pp_dict):
	"""Resolve whether a Production Plan document has had "Get Update" applied."""
	if not pp_dict:
		return False, False

	sales_order_mod = (
		pp_dict.get("sales_order_modification")
		if isinstance(pp_dict, dict)
		else getattr(pp_dict, "sales_order_modification", None)
	)
	bom_mod = (
		pp_dict.get("bom_modification")
		if isinstance(pp_dict, dict)
		else getattr(pp_dict, "bom_modification", None)
	)

	def _norm_mod(val):
		if val is None:
			return ""
		if isinstance(val, bool):
			return "1" if val else "0"
		return cstr(val).strip().upper()

	sales_order_mod_upper = _norm_mod(sales_order_mod)
	bom_mod_upper = _norm_mod(bom_mod)

	has_so_mod = sales_order_mod_upper in ("YES", "1", "TRUE")
	has_bom_mod = bom_mod_upper in ("YES", "1", "TRUE")
	has_any_mod = has_so_mod or has_bom_mod
	is_old_pp = not has_any_mod

	prod_plan_updated = (
		pp_dict.get("production_plan_updated")
		if isinstance(pp_dict, dict)
		else getattr(pp_dict, "production_plan_updated", 0)
	)
	pp_updated_flag = bool(cint(prod_plan_updated or 0))

	wo_updated = (
		pp_dict.get("work_order_updated")
		if isinstance(pp_dict, dict)
		else getattr(pp_dict, "work_order_updated", 0)
	)
	pp_wo_updated_flag = bool(cint(wo_updated or 0))

	if pp_updated_flag:
		pp_is_updated = True
	elif is_old_pp:
		pp_is_updated = True
	else:
		pp_is_updated = False

	return pp_is_updated, pp_wo_updated_flag


def _elapsed_from(start_time, now_dt=None):
	"""Return (days_elapsed, weeks_elapsed, hours_elapsed) from a timestamp.
	Optimized for high-throughput batch row processing."""
	if not start_time:
		return 0.0, 0.0, 0.0
	if now_dt is None:
		now_dt = now_datetime()

	if isinstance(start_time, dt):
		st = start_time
	elif isinstance(start_time, date_type):
		st = dt.combine(start_time, dt.min.time())
	elif isinstance(start_time, str):
		s = start_time.strip()
		if len(s) >= 19 and s[10] in (" ", "T"):
			try:
				st = dt(
					int(s[0:4]), int(s[5:7]), int(s[8:10]),
					int(s[11:13]), int(s[14:16]), int(s[17:19])
				)
			except (ValueError, TypeError):
				st = None
		elif len(s) == 10:
			try:
				st = dt(int(s[0:4]), int(s[5:7]), int(s[8:10]))
			except (ValueError, TypeError):
				st = None
		else:
			st = None

		if st is None:
			return 0.0, 0.0, 0.0
	else:
		return 0.0, 0.0, 0.0

	total_sec = (now_dt - st).total_seconds()
	days_elapsed = total_sec / 86400.0
	weeks_elapsed = days_elapsed / 7.0
	hours_elapsed = total_sec / 3600.0
	return days_elapsed, weeks_elapsed, hours_elapsed


def _format_elapsed_time_weeks(weeks):
	"""Format elapsed time in weeks for display."""
	if weeks < 0.14:
		return _("less than 1 day")
	elif weeks < 1:
		days = int(weeks * 7)
		return _("1 day") if days == 1 else _("{0} days").format(days)
	elif weeks < 2:
		return _("1 week")
	elif weeks < 3:
		return _("2 weeks")
	elif weeks < 4:
		return _("3 weeks")
	else:
		return _("{0} weeks").format(int(weeks))


def _compute_batch_row(so, omr, item, maps, now_dt=None):
	"""Compute one report row for a single CHANGED batch line inside one OMR."""
	so_name = so["name"] if isinstance(so, dict) else so.name
	get_so = so.get if isinstance(so, dict) else (lambda k, d=None: getattr(so, k, d))

	omr_name = omr.get("name") if isinstance(omr, dict) else omr.name
	omr_workflow_state = omr.get("workflow_state") if isinstance(omr, dict) else omr.workflow_state
	omr_creation = omr.get("creation") if isinstance(omr, dict) else omr.creation

	change_type = _line_change_type(item)
	effective_item = _effective_item(item)

	# ---- BMR: row's own linked BMR ----
	bmr_name = item.get("bom_update_request")
	bmr = dict(maps["bmr"][bmr_name]) if bmr_name and bmr_name in maps["bmr"] else None

	if change_type != "Item Replacement":
		bmr_status = "Not Required"
	elif not bmr:
		bmr_status = "Not Created"
	else:
		ds = cint(bmr.get("docstatus", 0))
		wf = bmr.get("workflow_state") or ""
		if ds == 2:
			bmr_status = "Rejected"
		elif ds == 0:
			bmr_status = "Draft"
		elif ds == 1 and wf == "Approved":
			bmr_status = "Updated"
		elif ds == 1 and wf in ("Rejected", "Cancelled"):
			bmr_status = "Rejected"
		else:
			bmr_status = "Pending"

	# ---- PP: resolved at BATCH + EFFECTIVE-ITEM + BATCH_NO level ----
	item_batch_no = item.get("batch_no")
	pp_dict = _resolve_pp_for_batch(so_name, effective_item, item_batch_no, maps)
	pp_exists = bool(pp_dict)
	pp_is_updated, pp_wo_updated_flag = _pp_update_flags(pp_dict)

	if change_type == "Item Replacement":
		if bmr_status != "Updated":
			pp_status = "Not Required"
		elif not pp_exists:
			pp_status = "Not Created"
		elif pp_is_updated:
			pp_status = "Updated"
		else:
			pp_status = "Pending"
	else:
		if not pp_exists:
			pp_status = "Not Created"
		elif pp_is_updated:
			pp_status = "Updated"
		else:
			pp_status = "Pending"

	# ---- WO: resolved via the batch's OWN resolved PP + effective item ----
	if pp_dict:
		k = (pp_dict["name"], effective_item)
		matched_wos = maps.get("wo_by_pp_item", {}).get(k, [])
	else:
		matched_wos = []

	if pp_status != "Updated":
		wo_status = "Not Created" if not matched_wos else "Not Started"
	elif not matched_wos:
		wo_status = "Not Created"
	else:
		wo_pending = any(
			cstr(w.get("modification_status") or "").strip().upper() in ("YES", "1", "TRUE")
			for w in matched_wos
		)
		wo_status = "Pending" if wo_pending else "Updated"

	# ---- pending / severity ----
	is_pending = False
	if bmr_status in ("Draft", "Pending", "Rejected"):
		is_pending = True
	elif bmr_status == "Not Created" and change_type == "Item Replacement":
		is_pending = True
	elif pp_status in ("Pending", "Not Created"):
		is_pending = True
	elif wo_status in ("Pending", "Not Created", "Not Started"):
		is_pending = True

	days_elapsed, weeks_elapsed, hours_elapsed = _elapsed_from(omr_creation, now_dt=now_dt)

	if is_pending:
		if weeks_elapsed >= 4:
			severity = "Critical"
		elif weeks_elapsed >= 2:
			severity = "High"
		else:
			severity = "Low"
	else:
		severity = "Low"

	# ---- pending-at / remarks ----
	pending_at = "—"
	if bmr_status == "Draft":
		pending_at = _("BMR Draft Submission")
	elif bmr_status == "Pending":
		pending_at = _("BMR Approval")
	elif bmr_status == "Rejected":
		pending_at = _("BMR Rejection Resolution")
	elif bmr_status == "Not Created" and change_type == "Item Replacement":
		pending_at = _("BMR Creation")
	elif pp_status == "Not Created":
		pending_at = _("Production Plan Creation")
	elif pp_status == "Pending":
		pending_at = _("Production Plan Update (Get Update)")
	elif wo_status in ("Not Created", "Not Started"):
		pending_at = _("Work Order Creation")
	elif wo_status == "Pending":
		pending_at = _("Work Order Update")

	if is_pending:
		elapsed_str = _format_elapsed_time_weeks(weeks_elapsed)
		remarks = _("Pending for {0} at {1}").format(elapsed_str, pending_at)
	else:
		remarks = _("Completed")

	# ---- stage rail ----
	stages = [{"key": "so", "label": "SO", "state": "completed"}]
	omr_state = "completed" if omr_workflow_state == "Approved" else "pending"
	stages.append({"key": "omr", "label": "OMR", "state": omr_state})

	if bmr_status in ("Not Created", "Not Required"):
		bmr_state = "not_required" if change_type != "Item Replacement" else "pending"
	elif bmr_status in ("Draft", "Pending"):
		bmr_state = "pending"
	elif bmr_status == "Rejected":
		bmr_state = "blocked"
	elif bmr_status == "Updated":
		bmr_state = "completed"
	else:
		bmr_state = "not_required"
	stages.append({"key": "bmr", "label": "BMR", "state": bmr_state})

	if pp_status in ("Not Created", "Not Required"):
		pp_state = "not_required" if pp_status == "Not Required" else "pending"
	elif pp_status == "Pending":
		pp_state = "blocked"
	elif pp_status == "Updated":
		pp_state = "completed"
	else:
		pp_state = "not_required"
	stages.append({"key": "pp", "label": "PP", "state": pp_state})

	if wo_status in ("Not Created", "Not Started"):
		wo_state = "not_required"
	elif wo_status == "Pending":
		wo_state = "blocked"
	elif wo_status == "Updated":
		wo_state = "completed"
	else:
		wo_state = "not_required"
	stages.append({"key": "wo", "label": "WO", "state": wo_state})

	stages.append({"key": "completed", "label": "Completed", "state": "completed" if not is_pending else "pending"})

	row_id = "{0}::{1}".format(so_name, item.get("name"))

	return {
		"row_id": row_id,
		"sales_order": so_name,
		"customer": get_so("customer"),
		"customer_name": get_so("customer_name"),
		"branch": get_so("branch"),
		"company": get_so("company"),
		"transaction_date": get_so("transaction_date"),
		"delivery_date": get_so("delivery_date"),
		"so_status": get_so("status"),
		"modified": get_so("modified"),
		"modified_by": get_so("modified_by"),

		"omr": omr_name,
		"omr_workflow_state": omr_workflow_state,
		"omr_creation": omr_creation,

		"batch_no": item.get("batch_no"),
		"tag_no": item.get("tag_no"),
		"item": item.get("item"),
		"rev_item": item.get("rev_item"),
		"effective_item": effective_item,
		"description": item.get("description"),
		"rev_description": item.get("rev_description"),
		"qty": flt(item.get("qty")),
		"rev_qty": item.get("rev_qty"),
		"line_status": item.get("line_status"),
		"rev_line_status": item.get("rev_line_status"),
		"change_type": change_type,

		"bmr": bmr,
		"bmr_status": bmr_status,
		"bmr_status_badge": _status_badge(bmr_status, "bmr"),

		"pp": pp_dict,
		"pp_status": pp_status,
		"pp_status_badge": _status_badge(pp_status, "pp"),

		"wo_list": matched_wos,
		"wo_status": wo_status,
		"wo_status_badge": _status_badge(wo_status, "wo"),

		"severity": severity,
		"severity_color": SEVERITY_COLOR[severity],
		"is_pending": is_pending,
		"stages": stages,
		"pending_at": pending_at,
		"remarks": remarks,
		"days_elapsed": round(days_elapsed, 1),
		"weeks_elapsed": round(weeks_elapsed, 1),
		"hours_elapsed": round(hours_elapsed, 1),
	}


def _compute_batch_rows_for_so(so, maps, now_dt=None):
	"""All changed batch rows (across every OMR) for a single Sales Order."""
	so_name = so["name"] if isinstance(so, dict) else so.name
	rows = []
	for omr in maps["omr_all"].get(so_name, []):
		omr_name = omr.get("name") if isinstance(omr, dict) else omr.name
		for item in maps["omr_items"].get(omr_name, []):
			if _check_item_has_changes(item):
				rows.append(_compute_batch_row(so, omr, item, maps, now_dt=now_dt))
	return rows


def _rows_cache_key(filters):
	"""Stable cache key for a given filter set."""
	payload = json.dumps(filters or {}, sort_keys=True, default=str)
	digest = hashlib.md5(payload.encode("utf-8")).hexdigest()
	return "ppwo_control_report:rows:{0}".format(digest)


def _bust_rows_cache():
	"""Invalidate every cached row-set."""
	try:
		frappe.cache().delete_keys("ppwo_control_report:rows:*")
	except Exception:
		pass
	if hasattr(frappe.local, "_ppwo_raw_rows_cache"):
		delattr(frappe.local, "_ppwo_raw_rows_cache")


def _get_raw_computed_rows(filters, limit=5000):
	"""Get all changed-batch rows across the filtered Sales Orders WITHOUT
	applying post-filters. Uses request-local caching and Redis caching."""
	cache_key = _rows_cache_key(filters)

	# Check frappe.local request-level cache first
	local_cache = getattr(frappe.local, "_ppwo_raw_rows_cache", None)
	if local_cache and cache_key in local_cache:
		return local_cache[cache_key][:limit]

	rows = frappe.cache().get_value(cache_key)

	if rows is None:
		sos = _get_filtered_sales_orders(filters)
		if not sos:
			rows = []
		else:
			so_names = [s["name"] if isinstance(s, dict) else s.name for s in sos]
			filtered_pp = filters.get("production_plan")
			maps = _fetch_maps(so_names, filtered_pp=filtered_pp)
			now_dt = now_datetime()

			rows = []
			for s in sos:
				rows.extend(_compute_batch_rows_for_so(s, maps, now_dt=now_dt))

		frappe.cache().set_value(cache_key, rows, expires_in_sec=ROWS_CACHE_TTL)

	if not hasattr(frappe.local, "_ppwo_raw_rows_cache"):
		frappe.local._ppwo_raw_rows_cache = {}
	frappe.local._ppwo_raw_rows_cache[cache_key] = rows

	return rows[:limit]


def _get_computed_rows(filters, limit=5000):
	"""Get all changed-batch rows across the filtered Sales Orders, then
	apply post-filters (pending_only, critical_only, status, priority)."""
	rows = _get_raw_computed_rows(filters, limit=limit)

	if cint(filters.get("pending_only")):
		rows = [r for r in rows if r["is_pending"]]

	if cint(filters.get("critical_only")):
		rows = [r for r in rows if r["severity"] == "Critical"]

	status_filter = filters.get("status")
	if status_filter and status_filter != "Any":
		if status_filter == "Fully Synced":
			rows = [r for r in rows if not r["is_pending"]]
		elif status_filter == "Pending Update":
			rows = [r for r in rows if r["is_pending"]]
		elif status_filter == "Waiting on Approval":
			rows = [r for r in rows if r["omr_workflow_state"] != "Approved"]

	priority_filter = filters.get("priority")
	if priority_filter and priority_filter != "Any":
		rows = [r for r in rows if r["severity"] == priority_filter]

	return rows[:limit]