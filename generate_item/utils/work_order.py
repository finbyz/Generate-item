import frappe
from frappe.utils import flt
import json
from frappe.utils import getdate, nowdate
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo

def before_save(doc, method=None):
    """Update current_qty before saving Work Order."""
    branch_warehouses = get_branch_raw_store_warehouses(doc.branch)

    for row in doc.required_items:
        row.current_qty = get_current_qty_from_bin(
            item_code=row.item_code,
            source_warehouse=row.source_warehouse or doc.source_warehouse,
            target_warehouse=doc.fg_warehouse,  
        )
        

        row.on_hand_qty = get_branch_on_hand_qty(
            row.item_code,
            tuple(branch_warehouses)
        )

def get_current_qty_from_bin(item_code, source_warehouse=None, target_warehouse=None):
    """Sum actual_qty from Bin for source and target warehouses."""

    if not item_code:
        return 0

    warehouses = {
        wh for wh in (source_warehouse, target_warehouse) if wh
    }

    if not warehouses:
        return 0

    qty = frappe.db.sql(
        """
        SELECT COALESCE(SUM(actual_qty), 0)
        FROM `tabBin`
        WHERE item_code = %s
          AND warehouse IN %s
        """,
        (item_code, tuple(warehouses)),
    )[0][0]

    return qty or 0


def get_branch_raw_store_warehouses(branch):
    if not branch:
        return []

    return frappe.get_all(
        "Warehouse",
        filters={
            "branch": branch,
            "disabled": 0,
        },
        or_filters={
            "raw_material_warehouse": 1,
            "store_warehouse": 1,
        },
        pluck="name",
    )

def before_insert(doc, method=None):

    logger = frappe.logger("generate_item")
    try:
        logger.info(f"Work Order before_insert: name={getattr(doc, 'name', None)}, sales_order={getattr(doc, 'sales_order', None)}, bom_no={getattr(doc, 'bom_no', None)}")

        if getattr(doc, 'sales_order', None):
            sales_order = frappe.get_doc('Sales Order', doc.sales_order)
            logger.info(f"Fetched Sales Order {sales_order.name} with {len(sales_order.items or [])} items")

            for item in (getattr(doc, 'required_items', []) or []):
                if hasattr(item, 'sales_order'):
                    item.sales_order = doc.sales_order

            batch_no = None
            branch = None

            sales_order_item_name = getattr(doc, 'sales_order_item', None)
            if sales_order_item_name:
                soi = frappe.get_all(
                    'Sales Order Item',
                    filters={'name': sales_order_item_name, 'parent': sales_order.name},
                    fields=['name', 'item_code', 'custom_batch_no', 'bom_no', 'idx', 'branch'],
                    limit=1,
                )
                if soi:
                    batch_no = soi[0].get('custom_batch_no')
                    branch = soi[0].get('branch')
                    logger.info(f"Batch from exact Sales Order Item {sales_order_item_name}: {batch_no}")

            if not batch_no:
                production_item = getattr(doc, 'production_item', None) or getattr(doc, 'item_code', None)
                bom_no = getattr(doc, 'bom_no', None)
                if production_item:
                    soi_filters = {'parent': sales_order.name, 'item_code': production_item}
                    soi_fields = ['name', 'custom_batch_no', 'bom_no', 'idx', 'branch']
                    candidates = frappe.get_all('Sales Order Item', filters=soi_filters, fields=soi_fields, order_by='idx asc')
                    if candidates:
                        chosen = None
                        if bom_no:
                            for c in candidates:
                                if c.get('bom_no') == bom_no and c.get('custom_batch_no'):
                                    chosen = c
                                    break
                        if not chosen:
                            chosen = next((c for c in candidates if c.get('custom_batch_no')), candidates[0])
                        batch_no = chosen.get('custom_batch_no')
                        if not branch:
                            branch = chosen.get('branch')
                        logger.info(f"Batch from SO item by item_code/BOM: {batch_no}")

            if batch_no:
                doc.custom_batch_no = batch_no
                for item in (getattr(doc, 'required_items', []) or []):
                    item.custom_batch_no = batch_no
                for item in (getattr(doc, 'items', []) or []):
                    item.custom_batch_no = batch_no
            else:
                logger.info("Could not resolve custom_batch_no from Sales Order context")

            if branch:
                try:
                    for child in (getattr(doc, 'required_items', []) or []):
                        if hasattr(child, 'branch') and not getattr(child, 'branch', None):
                            child.branch = branch
                    for child in (getattr(doc, 'items', []) or []):
                        if hasattr(child, 'branch') and not getattr(child, 'branch', None):
                            child.branch = branch
                    logger.info(f"Applied branch {branch} to Work Order child rows where missing")
                except Exception as set_branch_err:
                    logger.error(f"Failed to set child branch: {set_branch_err}")

        if getattr(doc, 'bom_no', None):
            bom = frappe.get_doc("BOM", doc.bom_no)
            logger.info(f"Fetched BOM {bom.name} with {len(bom.items or [])} items")

            doc.custom_ga_drawing_no = getattr(bom, 'custom_ga_drawing_no', None)
            doc.custom_ga_drawing_rev_no = getattr(bom, 'custom_ga_drawing_rev_no', None)

            bom_items_map = {d.item_code: d for d in (bom.items or []) if getattr(d, 'item_code', None)}

            children = (getattr(doc, 'required_items', []) or []) + (getattr(doc, 'items', []) or [])
            for child in children:
                bom_item = bom_items_map.get(getattr(child, 'item_code', None))
                if not bom_item:
                    continue
                child.custom_drawing_no = getattr(bom_item, 'custom_drawing_no', None)
                child.custom_pattern_drawing_no = getattr(bom_item, 'custom_pattern_drawing_no', None)
                child.custom_purchase_specification_no = getattr(bom_item, 'custom_purchase_specification_no', None)
                child.custom_drawing_rev_no = getattr(bom_item, 'custom_drawing_rev_no', None)
                child.custom_pattern_drawing_rev_no = getattr(bom_item, 'custom_pattern_drawing_rev_no', None)
                child.custom_purchase_specification_rev_no = getattr(bom_item, 'custom_purchase_specification_rev_no', None)
                child.custom_batch_no = getattr(bom_item, 'custom_batch_no', None) or getattr(child, 'custom_batch_no', None)

    except Exception as e:
        logger.error(f"Error in work_order.before_insert: {e}")
        frappe.log_error(frappe.get_traceback(), "work_order.before_insert")


def on_trash(doc, method=None):
    logger = frappe.logger("generate_item")
    try:
        logger.info(f"Work Order on_trash: name={doc.name}, production_plan={getattr(doc, 'production_plan', None)}")

        production_plan_name = getattr(doc, 'production_plan', None)

        # Fallback 1: search via Production Plan Item
        if not production_plan_name:
            linked = frappe.get_all(
                'Production Plan Item',
                filters={'work_order': doc.name},
                fields=['name', 'parent'],
                limit=1
            )
            if linked:
                production_plan_name = linked[0].get('parent')

        # Fallback 2: search via Sub Assembly Item
        if not production_plan_name:
            linked_sa = frappe.get_all(
                'Production Plan Sub Assembly Item',
                filters={'work_order': doc.name},
                fields=['name', 'parent'],
                limit=1
            )
            if linked_sa:
                production_plan_name = linked_sa[0].get('parent')

        if not production_plan_name:
            logger.info(f"No Production Plan linked to Work Order {doc.name}, skipping reset")
            return

        logger.info(f"Resetting Production Plan {production_plan_name} for deleted WO {doc.name}")

        # ── Reset Assembly Items (po_items) ──────────────────────────────
        po_items = frappe.get_all(
            'Production Plan Item',
            filters={
                'parent': production_plan_name,
                'work_order': doc.name
            },
            fields=['name', 'ordered_qty', 'planned_qty', 'produced_qty']
        )

        for item in po_items:
            # Subtract only this WO's qty; never go below 0
            new_ordered_qty = max(0, flt(item.ordered_qty) - flt(doc.qty))
            frappe.db.set_value(
                'Production Plan Item',
                item.name,
                {
                    'work_order': None,
                    'ordered_qty': new_ordered_qty
                }
            )
            logger.info(
                f"Reset po_item {item.name}: work_order cleared, "
                f"ordered_qty {item.ordered_qty} -> {new_ordered_qty}"
            )

        # ── Reset Sub Assembly Items ──────────────────────────────────────
        sa_items = frappe.get_all(
            'Production Plan Sub Assembly Item',
            filters={
                'parent': production_plan_name,
                'work_order': doc.name
            },
            fields=['name', 'qty']
        )

        for sa_item in sa_items:
            update_values = {'work_order': None}

            # Check if wo_created field exists before setting it
            meta = frappe.get_meta('Production Plan Sub Assembly Item')
            field_names = [f.fieldname for f in meta.fields]
            if 'wo_created' in field_names:
                update_values['wo_created'] = 0

            frappe.db.set_value(
                'Production Plan Sub Assembly Item',
                sa_item.name,
                update_values
            )
            logger.info(f"Reset sub_assembly_item {sa_item.name}: {update_values}")

        # ── Recalculate and fix Production Plan status ────────────────────
        _reset_production_plan_status(production_plan_name, logger)

    except Exception as e:
        logger.error(f"Error in work_order.on_trash: {e}")
        frappe.log_error(frappe.get_traceback(), "work_order.on_trash")


def _reset_production_plan_status(production_plan_name, logger=None):
    if not logger:
        logger = frappe.logger("generate_item")

    try:
        pp = frappe.db.get_value(
            'Production Plan',
            production_plan_name,
            ['status', 'docstatus'],
            as_dict=True
        )

        if not pp or pp.docstatus != 1:
            return

        # Check if ANY work order still exists for this production plan
        existing_wo_count = frappe.db.count(
            'Work Order',
            filters={'production_plan': production_plan_name}
        )

        # If no WOs remain, OR fewer WOs than po_items+sub_assembly_items, show button
        po_item_count = frappe.db.count(
            'Production Plan Item',
            filters={'parent': production_plan_name}
        )
        sub_item_count = frappe.db.count(
            'Production Plan Sub Assembly Item',
            filters={'parent': production_plan_name}
        )

        total_needed = po_item_count + sub_item_count
        has_pending = existing_wo_count < total_needed

        logger.info(
            f"PP {production_plan_name}: existing_wo={existing_wo_count}, "
            f"needed={total_needed}, has_pending={has_pending}"
        )

        if has_pending:
            frappe.db.set_value(
                'Production Plan',
                production_plan_name,
                'status',
                'Material Requested'
            )

            logger.info(f"Production Plan {production_plan_name} status -> 'Material Requested'")

    except Exception as e:
        logger.error(f"Error in _reset_production_plan_status: {e}")
        frappe.log_error(frappe.get_traceback(), "_reset_production_plan_status")



from frappe import _
from frappe.utils import flt
from erpnext.manufacturing.doctype.bom.bom import get_bom_items_as_dict

from erpnext.manufacturing.doctype.work_order.work_order import stop_unstop


@frappe.whitelist()
def get_update_for_work_order(docname):
    wo = frappe.get_doc("Work Order", docname)
    _block_if_pp_has_pending_modifications(wo.production_plan, wo.name)
    result = _update_single_work_order(wo)

    return result


@frappe.whitelist()
def clear_work_order_updated(docname):
    frappe.db.set_value(
        "Production Plan",
        docname,
        "work_order_updated",
        0,
        update_modified=True
    )



@frappe.whitelist()
def get_update_for_production_plan(docname):
    _block_if_pp_has_pending_modifications(docname)
    """
    Run the same BOM sync logic for every Work Order linked to this
    Production Plan. Each WO is updated independently — if one is
    blocked (e.g. already Started) or errors out, the rest still proceed.
    """
    wo_names = frappe.get_all(
        "Work Order",
        filters={"production_plan": docname, "docstatus": ["!=", 2]},
        pluck="name",
    )

    if not wo_names:
        frappe.throw(_("No active Work Orders found against this Production Plan."))

    results = []

    for wo_name in wo_names:
        # Savepoint so one WO's failure doesn't roll back the ones
        # already updated in this same request.
        savepoint = f"wo_update_{wo_name}"
        frappe.db.savepoint(savepoint)
        try:
            wo = frappe.get_doc("Work Order", wo_name)
            res = _update_single_work_order(wo)
            results.append({"work_order": wo_name, "status": "success", "message": res.get("message")})
        except Exception as e:
            frappe.db.rollback(save_point=savepoint)
            results.append({"work_order": wo_name, "status": "failed", "message": str(e)})
            frappe.log_error(
                title=_("Production Plan WO update failed"),
                message=frappe.get_traceback(),
            )



    return {
        "total": len(results),
        "succeeded": len([r for r in results if r["status"] == "success"]),
        "failed": len([r for r in results if r["status"] == "failed"]),
        "details": results,
    }
    
def _update_single_work_order(wo):
    """
    Core update logic for one Work Order. Shared by the single-WO
    endpoint and the Production Plan bulk endpoint.
    Raises frappe.ValidationError (via frappe.throw) for blocked WOs —
    caller decides whether that aborts everything (single) or is just
    recorded and skipped (bulk).
    """
    allowed_statuses = ("Draft", "Not Started")

    if wo.docstatus == 2:
        frappe.throw(_("Cannot update a cancelled Work Order ({0}).").format(wo.name))

    if wo.status not in allowed_statuses:
        frappe.throw(
            _("Work Order {0} cannot be updated because it is already in <b>{1}</b> status.")
            .format(wo.name, wo.status)
        )

    if not wo.bom_no:
        frappe.throw(_("Work Order {0} does not have a BOM linked.").format(wo.name))

    bom = frappe.get_doc("BOM", wo.bom_no)

    if bom.item and bom.item != wo.production_item:
        wo.db_set("production_item", bom.item, update_modified=False)
        wo.db_set("item_name", bom.item_name, update_modified=False)
        wo.db_set("description", bom.description, update_modified=False)

    _sync_required_items_from_bom(wo, bom)
    wo.reload()
    # stop_unstop(wo.name,"Stopped")
    # stop_unstop(wo.name,"Resumed")
    wo.update_status()
    wo.update_planned_qty()



    wo.db_set("modification_status", "No", update_modified=False)

    return {"success": True, "message": _("Work Order {0} updated").format(wo.name)}

def _sync_required_items_from_bom(wo, bom):
    """
    Sync required_items with the current BOM, without losing history.
    Works for draft and submitted Work Orders.
    """
    bom_items = get_bom_items_as_dict(
        bom.name,
        wo.company,
        qty=wo.qty,
        fetch_exploded=wo.use_multi_level_bom,
        fetch_secondary_items=0,
    )
    bom_item_codes = set(bom_items.keys())
    existing_rows = {row.item_code: row for row in wo.required_items}

    rows_to_keep = []
    blocked_removals = []

    for item_code, row in existing_rows.items():
        if item_code in bom_item_codes:
            bi = bom_items[item_code]
            row.required_qty = bi.qty
            row.rate = bi.rate
            row.amount = flt(bi.qty) * flt(bi.rate)
            rows_to_keep.append(row)
        elif flt(row.transferred_qty) or flt(row.consumed_qty):
            blocked_removals.append(row.item_code)
            rows_to_keep.append(row)

    if blocked_removals:
        frappe.msgprint(
            _("Work Order {0}: items {1} were removed from the BOM but kept "
              "because material has already been transferred or consumed against them.")
            .format(wo.name, ", ".join(blocked_removals)),
            indicator="orange",
            alert=True,
        )

    for item_code, bi in bom_items.items():
        if item_code not in existing_rows:
            rows_to_keep.append(frappe._dict({
                "item_code": item_code,
                "item_name": bi.item_name,
                "description": bi.description,
                "stock_uom": bi.stock_uom,
                "required_qty": bi.qty,
                "rate": bi.rate,
                "amount": flt(bi.qty) * flt(bi.rate),
                "source_warehouse": bi.source_warehouse or wo.source_warehouse,
                "allow_alternative_item": bi.allow_alternative_item,
                "include_item_in_manufacturing": bi.include_item_in_manufacturing,
                "branch": wo.get("branch"),
                "sales_order": wo.get("sales_order"),
                "custom_batch_no": wo.get("custom_batch_no"),
            }))

    wo.set("required_items", rows_to_keep)
    wo.flags.ignore_validate_update_after_submit = True
    wo.save(ignore_permissions=True)

def remove_modification_task_link(doc, method=None):
    """
    Before a draft Work Order is trashed, unlink it from any
    Modification Task referencing it, so the Dynamic Link check
    doesn't block deletion.
    """
    # Safety: only auto-unlink for Draft WOs. Anything further along,
    # fail loud instead of silently orphaning a task.
    if doc.status != "Draft":
        return

    linked_tasks = frappe.get_all(
        "Modification Task",
        filters={
            "reference_doctype": "Work Order",
            "reference_document_name": doc.name,
        },
        pluck="name",
    )

    for task_name in linked_tasks:
        # Modification Task is submittable (docstatus=1), so a normal
        # doc.save() would fail on non-allow-on-submit fields.
        # db_set bypasses validate/on_update and works on submitted docs.
        note = (
            f"\n\n---\n[System] Reference Work Order **{doc.name}** was "
            f"deleted on {frappe.utils.now()}. Link removed automatically."
        )
        existing_desc = frappe.db.get_value("Modification Task", task_name, "description") or ""

        frappe.db.set_value(
            "Modification Task",
            task_name,
            {
                "reference_doctype": None,
                "reference_document_name": None,
                "description": existing_desc + note,
            },
            update_modified=True,
        )



def get_branch_warehouses(branch):
    """Return list of all warehouse names belonging to a branch."""
    if not branch:
        return []
    return frappe.get_all(
        "Warehouse",
        filters={"branch": branch, "disabled": 0},
        pluck="name"
    )
 
 
def get_branch_on_hand_qty(item_code, warehouses):
    """Sum actual_qty (Bin) for an item across all given warehouses."""
    if not item_code or not warehouses:
        return 0
 
    total = frappe.db.sql(
        """
        select sum(actual_qty)
        from `tabBin`
        where item_code = %s
        and warehouse in %s
        """,
        (item_code, warehouses),
    )[0][0]
 
    return total or 0

def get_current_qty(item_code, source_warehouse, target_warehouse):
    """Sum actual_qty (Bin) for an item across source + target warehouse (dedup if same)."""
    if not item_code:
        return 0

    warehouses = set()
    if source_warehouse:
        warehouses.add(source_warehouse)
    if target_warehouse:
        warehouses.add(target_warehouse)

    if not warehouses:
        return 0

    total = frappe.db.sql(
        """
        select sum(actual_qty)
        from `tabBin`
        where item_code = %s
        and warehouse in %s
        """,
        (item_code, tuple(warehouses)),
    )[0][0]

    return total or 0
 
 
@frappe.whitelist()
def export_work_orders(work_orders):
    if isinstance(work_orders, str):
        work_orders = json.loads(work_orders)
 
    wb = Workbook()
    ws = wb.active
    ws.title = "Work Orders"
 
    headers = [
        "WO No",
        "Branch",
        "FG Item Code",
        "BOM No",
        "Batch No",
        "Item Code",
        "Item Description",
        "Qty Issued",
        "Current Qty",   
        "On Hand Qty",
        "Balance Qty",
        "UOM",
        "Source WH",
        "Bin No",
        "Target WH",
        "Drawing No",
        "Drawing Rev No",
    ]
 
    ws.append(headers)
 
    # Header Style
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
 
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
 
    # cache: branch -> list of warehouses (avoid repeat queries)
    branch_warehouse_cache = {}
 
    for wo_name in work_orders:
        wo = frappe.get_doc("Work Order", wo_name)
 
        # NOTE: adjust these fieldnames if your Work Order doctype
        # uses different fieldnames for branch / batch_no
        branch = getattr(wo, "branch", "") or ""
        fg_item_code = wo.production_item or ""
        bom_no = wo.bom_no or ""
        custom_batch_no = getattr(wo, "custom_batch_no", "") or ""
 
        if branch not in branch_warehouse_cache:
            branch_warehouse_cache[branch] = get_branch_warehouses(branch)
 
        branch_warehouses = branch_warehouse_cache[branch]
 
        if not wo.get("required_items"):
            # still emit a row with header-level info if there are no items
            row = [
                wo.name,
                branch,
                fg_item_code,
                bom_no,
                custom_batch_no,
                "", "", "", "", "", "", "", "", "", "", "", "", 
            ]
            ws.append(row)
            continue
 
        for item in wo.required_items:
            on_hand_qty = get_branch_on_hand_qty(item.item_code, branch_warehouses)
            current_qty = get_current_qty(
                item.item_code,
                item.source_warehouse,
                wo.fg_warehouse
            )
 
            row = [
                wo.name,
                branch,
                fg_item_code,
                bom_no,
                custom_batch_no,
                item.item_code or "",
                item.description or "",
                item.transferred_qty or 0,
                current_qty,    
                on_hand_qty,
                item.available_qty_at_wip_warehouse or 0,
                item.stock_uom or "",
                item.source_warehouse or "",
                "",
                wo.fg_warehouse or "",
                item.custom_drawing_no or "",
                item.custom_drawing_rev_no or "",
            ]
            ws.append(row)
 
    # Apply border and alignment
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
 
    # Column widths
    widths = {
                "A": 20,  # WO No
                "B": 15,  # Branch
                "C": 42,  # FG Item Code
                "D": 42,  # BOM No
                "E": 15,  # Batch No
                "F": 42,  # Item Code
                "G": 35,  # Item Description
                "H": 12,  # Qty Issued
                "I": 12,  # Current Qty
                "J": 12,  # On Hand Qty
                "K": 12,  # Balance Qty
                "L": 10,  # UOM
                "M": 27,  # Source WH
                "N": 15,  # Bin No
                "O": 27,  # Target WH
                "P": 20,  # Drawing No
                "Q": 10,  # Drawing Rev No
            }
 
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
 
    # Format Qty columns
    for col_letter in ["H", "I", "J", "K"]:
        for cell in ws[col_letter][1:]:
            cell.number_format = '#,##0.00'
 
    # Freeze Header Row
    ws.freeze_panes = "A2"
 
    # Add Filter
    # ws.auto_filter.ref = ws.dimensions
 
    # Create Table
    table = Table(
        displayName="WorkOrderTable",
        ref=f"A1:Q{ws.max_row}"
    )
 
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
 
    table.tableStyleInfo = style
    ws.add_table(table)
 
    # Save File
    file_name = "Work_Order_Export.xlsx"
    file_path = f"/tmp/{file_name}"
 
    wb.save(file_path)
 
    with open(file_path, "rb") as f:
        file_doc = frappe.get_doc({
            "doctype": "File",
            "file_name": file_name,
            "is_private": 0,
            "content": f.read()
        })
        file_doc.save(ignore_permissions=True)
 
    return file_doc.file_url

# =============================================================================
# Production Plan modification gating for Work Order updates
# =============================================================================

def _block_if_pp_has_pending_modifications(production_plan, wo_name=None):
    """
    Check whether the linked Production Plan has a pending SO or BOM
    modification that must be absorbed first. Raises frappe.throw if so.

    Called before any Work Order update (single or bulk) to enforce the
    sequencing rule: Production Plan must absorb its own changes before
    Work Orders can be updated.
    """
    if not production_plan:
        return

    flags = frappe.db.get_value(
        "Production Plan",
        production_plan,
        ["sales_order_modification", "bom_modification"],
        as_dict=True,
    )

    if not flags:
        return

    blocked_by = []
    if flags.sales_order_modification == "YES":
        blocked_by.append("Sales Order Modification")
    if flags.bom_modification == "YES":
        blocked_by.append("BOM Modification")

    if blocked_by:
        wo_ref = wo_name or _("this Work Order")
        frappe.throw(
            _(
                "Cannot update {0}: the linked Production Plan <b>{1}</b> "
                "still has pending changes ({2}) that must be absorbed first. "
                "Please run <b>Get Update</b> on the Production Plan before "
                "updating Work Orders."
            ).format(wo_ref, production_plan, ", ".join(blocked_by)),
            title=_("Production Plan Not Updated"),
        )

